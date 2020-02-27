#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: gems_compare_user_rights
# Author: Samuel Bramm
# Date: 2018-08-09
# Description: Provides a delta comparison between two GEMS user right excel sheets
# Version: v1.0

from openpyxl import *
from Tkinter import *
from ttk import *
from tkFileDialog import askopenfilename
import tkMessageBox
import os
import time




root = Tk()
old_file_path = ""
new_file_path = ""
old_set = IntVar()
new_set = IntVar()


def write_excel(user_changes, deleted_users, added_users):
    global old_file_path
    out_file_path = os.path.dirname(old_file_path) + "/gems_user_changes_{}.xlsx".format(time.strftime("%Y_%m_%d"))
    book = Workbook()
    sheet = book.active
    sheet.title = "added and deleted users"
    sheet.append(["UID", "change"])
    for d in deleted_users:
        sheet.append([d, "deleted"])
    for a in added_users:
        sheet.append([a, "added"])
    book.create_sheet("changed users")
    sheet = book.get_sheet_by_name("changed users")
    sheet.append(['User-ID', 'Parameter', 'old value', 'new value'])
    for change in user_changes:
        sheet.append(change)
    try:
        book.save(out_file_path)
    except IOError:
        tkMessageBox.showerror("I/O Error",
                               "The file %s is not writeable. Please close all open instances of Excel and try again" % out_file_path)
        exit(0)

    return out_file_path


def compare_body(old_rights, new_rights, mapping, header):
    # compare body of both excel files
    deleted_users = []
    added_users = []
    old_users = [x[0] for x in old_rights]
    new_users = [x[0] for x in new_rights]
    user_changes = []

    # check if users were deleted or added
    for old in old_users:
        if old not in new_users and old not in deleted_users:
            deleted_users.append(old)
    for new in new_users:
        if new not in old_users and new not in added_users:
            added_users.append(new)
    # print mapping
    # check rights for not deleted users
    remaining_users = [x for x in old_users if x not in deleted_users]
    for user in remaining_users:
        old_right = [x for x in old_rights if x[0] == user][0]
        new_right = [x for x in new_rights if x[0] == user][0]

        # print old_right
        # print new_right
        for i in range(len(old_right)):
            if old_right[i] != new_right[mapping[i][1]]:
                # print "missmatch", user, header[i], old_right[i], new_right[mapping[i][1]]
                user_changes.append([user, header[i], old_right[i], new_right[mapping[i][1]]])

    return user_changes, deleted_users, added_users


def compare_headers(old_header, new_header):
    # compare headers and find matching parameters
    # print "old", len(old_header), old_header
    # print "new", len(new_header), new_header
    mapping = []
    if len(old_header) != len(new_header):
        tkMessageBox.showerror("Error", "The length of the headers in the two given files does not match. Aborting!")

    else:
        for i, entry in enumerate(old_header):
            try:
                new = new_header.index(entry)
                mapping.append([i, new])
            except:
                tkMessageBox.showerror("Error",
                                       "The header entry {} could not be found in the new file. Aborting!".format(
                                           entry))

    return mapping


def parse_excel(file_path):
    # parse given excel file and return header and body
    file_content = []
    try:
        book = load_workbook(file_path)
        sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
        for row in sheet.iter_rows():
            tmp_row = []
            for cell in row:
                tmp_row.append(cell.value)
            file_content.append(tmp_row)
    except:
        tkMessageBox.showerror("I/O Error",
                               "The file %s is not accessible. Please close all open versions and try again" % file_path)
    return file_content[:1][0], file_content[1:]


def compare_files(compare_button):
    # main comparison logic
    global root
    global old_file_path
    global new_file_path
    compare_button.config(state=DISABLED)
    Label(root, text="Reading old Excel file.").grid(row=3, column=0, padx=10, pady=10)
    root.update()
    old_header, old_rights = parse_excel(old_file_path)
    Label(root, text="Reading new Excel file.").grid(row=4, column=0, padx=10, pady=10)
    root.update()
    new_header, new_rights = parse_excel(new_file_path)
    mapping = compare_headers(old_header, new_header)

    Label(root, text="Comparing Excel files.").grid(row=5, column=0, padx=10, pady=10)
    root.update()
    user_changes, deleted_users, added_users = compare_body(old_rights, new_rights, mapping, old_header)

    Label(root, text="Writing Output file.").grid(row=6, column=0, padx=10, pady=10)
    root.update()
    out_file_path = write_excel(user_changes, deleted_users, added_users)
    if tkMessageBox.askyesno("Sucess",
                             "The Analysis is finished and the output file has been written to {}. Do you want to quit?".format(
                                 out_file_path)):
        root.quit()
    return None


def select_file(time_param):
    # read file path and store it to global variables
    file_path = askopenfilename(title="Select thr {} user right file".format(time_param),
                                filetypes=[("Excel file", "*.xlsx")])
    if time_param == "old":
        global old_file_path
        global old_set
        old_file_path = file_path
        old_set.set(1)
    else:
        global new_file_path
        global new_set
        new_file_path = file_path
        new_set.set(1)
    return None


def main():
    # display main menu and selections
    global root
    global old_set
    global new_set
    root.title("GEMS User rights comparison")
    Label(root, text="1. Select the old GEMS user right excel file").grid(row=0, column=0, padx=10, pady=10)
    old_button = Button(root, text="select", command=lambda: select_file("old"))
    old_button.grid(row=0, column=1, padx=10, pady=10)
    Label(root, text="2. Select the new GEMS user right excel file").grid(row=1, column=0, padx=10, pady=10)
    new_button = Button(root, text="select", command=lambda: select_file("new"))
    new_button.grid(row=1, column=1, padx=10, pady=10)
    root.wait_variable(old_set)
    old_button.config(state=DISABLED)
    root.wait_variable(new_set)
    new_button.config(state=DISABLED)
    Label(root, text="3. Press the compare button and wait").grid(row=2, column=0, padx=10, pady=10)
    compare_button = Button(root, text="compare", command=lambda: compare_files(compare_button))
    compare_button.grid(row=2, column=1, padx=10, pady=10)

    root.mainloop()
    exit(0)


if __name__ == '__main__':
    main()
