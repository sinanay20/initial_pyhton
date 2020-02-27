#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: Vega_I156_Dealer_Comparison
# Author: Benjamin Schmelcher
# Date: 20190206
# Description: Reads DB Extract file from icon and from Vega and compares (missing in icon/vega)
# Version: 1.0.0
import csv
import os
import tkMessageBox
from Tkinter import Tk, Text, DISABLED, END, W, IntVar
from tkFileDialog import askopenfilename
from ttk import Label, Button

from openpyxl import load_workbook, Workbook

#Statement for all tenants except china: Select * From icon.PARTNER_PARTNER where PARTNERTYPE ='workshop' and STATE='active' and TENANTID in ('51331BE', '53137FR', '56130PT', '77530BR')
#Statement for china: Select * From icon.PARTNER_PARTNER where PARTNERTYPE ='workshop' and STATE='active' and TENANTID = '81930CN'


root = Tk()



icon_path = ""
icon_path_set = IntVar()
vega_path = ""
vega_path_set = IntVar()
parse_started = IntVar()
outpath = ""


icon_header = []
vega_header = []

icon_list = []
vega_list = []

def choose_icon_file(button):
    global icon_path
    global icon_path_set

    icon_path = askopenfilename(parent=root, title="Choose iCON DB Extract", filetypes=[("CSV", "*.csv")])
    if icon_path != "":
        button.config(state=DISABLED)
        icon_path_set.set(1)
def choose_vega_file(button):
    global vega_path
    global vega_path_set

    vega_path = askopenfilename(parent=root, title="Choose VEGA Dealer File", filetypes=[("Excel", "*.xlsx")])
    if vega_path != "":
        button.config(state=DISABLED)
        vega_path_set.set(1)

def parse_files():
    global parse_started
    global icon_path
    global root
    global icon_list
    global vega_list
    global icon_header
    global vega_header
    global outpath

    #parse_started.set(1)
    with open(icon_path, 'rb') as icon:
        icon_reader = csv.reader(icon, delimiter=";")
        for idx, line in enumerate(icon_reader):
            root.update()
            root.update_idletasks()
            templine = []
            for entry in line:
                templine.append(entry.decode("utf-8"))
            if idx>0:
                icon_list.append(templine)
            else:
                icon_header = templine

    vega_book = load_workbook(vega_path)
    for sheet in vega_book.sheetnames:
        activesheet = vega_book.get_sheet_by_name(sheet)
        for idx, line in enumerate(activesheet.iter_rows()):
            templine = []
            for cell in line:
                templine.append(cell.value)
            if idx>0:
                if line[4].value.upper() == "RELEASED":
                    vega_list.append(templine)
            else:
                vega_header = templine

    found_icon_list = []
    found_vega_list = []
    missing_list = []

    #compare which entries exist in the db
    for vega in vega_list:
        found = False
        for icon in icon_list:
            if vega[0] == icon[29].zfill(5):
                found_icon_list.append(icon)
                if not found:
                    found_vega_list.append(vega)
                found = True

        if not found:
            missing_list.append(vega)

    print "found icon: " + str(len(found_icon_list))
    print "found vega: " + str(len(found_vega_list))
    print "missing vega: " + str(len(missing_list))


    #write found icon and missing vega dealers to workbook
    outpath = os.path.dirname(icon_path)+"/comparison.xlsx"

    out_book = Workbook()



    activesheet = out_book.active
    activesheet.title = "Missing in iCON"

    activesheet.append(vega_header)
    for missing in missing_list:
        activesheet.append(missing)

    try:
        out_book.save(outpath)
    except:
        tkMessageBox.showerror("Error", "Error writing to file.\nPlease make sure you have not opened the file {}."
                                        "\nAborting script.".format(outpath))
        exit(-1)


    #tenant specific excel files:
    out_book_belgium = Workbook()
    out_book_france = Workbook()
    out_book_china = Workbook()
    out_book_portugal = Workbook()
    out_book_brazil = Workbook()

    tenant_workbooks = [out_book_belgium, out_book_france, out_book_china, out_book_portugal, out_book_brazil]
    #icon_tenantstrings = ["51331BE", "53137FR", "81930CN", "56130PT", "77530BR"] # order has to be the same like in tenant_workbooks
    vega_tenantstrings = ["51331 - MBBEL", "53137 - MBF", "81930 - MBCL", "56130 - MBP", "77532 - MBBRAS"]# order has to be the same like in tenant_workbooks


    for idx, workbook in enumerate(tenant_workbooks):
        activesheet = workbook.active
        activesheet.title="Missing in iCON"
        activesheet.append(vega_header)
        for missing in missing_list:
            if missing[2] == vega_tenantstrings[idx]:
                activesheet.append(missing)

        try:
            outpath = os.path.dirname(icon_path)+"/comparison"+vega_tenantstrings[idx]+".xlsx"
            workbook.save(outpath)
        except:
            tkMessageBox.showerror("Error", "Error writing to file.\nPlease make sure you have not opened the file {}."
                                            "\nAborting script.".format(outpath))
            exit(-1)






def main():
    global icon_path_set
    global vega_path_set
    global parse_started
    global root
    global outpath

    root.title("Vega I156 Dealer/Workshop Comparison")


    icon_label = Label(root, text="1. Choose the iCON DB extract file containing all tenants.\n(Copy china data to other markets)")
    icon_label.grid(row=0, column=0, padx=10, pady=10, sticky=W)
    icon_button = Button(root, text="select", command=lambda: choose_icon_file(icon_button))
    icon_button.grid(row=0, column=1, padx=10, pady=10)
    root.wait_variable(icon_path_set)

    vega_label = Label(root, text="2. Choose the VEGA dealers/workshop file.")
    vega_label.grid(row=1, column=0, padx=10, pady=10, sticky=W)
    vega_button = Button(root, text="select", command=lambda: choose_vega_file(vega_button))
    vega_button.grid(row=1, column=1, padx=10, pady=10)
    root.wait_variable(vega_path_set)

    parse_label = Label(root, text="3. Click the parse button.")
    parse_label.grid(row=2, column=0, padx=10, pady=10, sticky=W)
    parse_button = Button(root, text="parse", command=lambda: parse_started.set(1))
    parse_button.grid(row=2, column=1, padx=10, pady=10)

    root.wait_variable(parse_started)

    progress_label = Label(root, text="Parsing files. Please wait...")
    progress_label.grid(row=3, column=0, columnspan=2, padx=10, pady=10)
    parse_files()
    progress_label.config(text="Parsing done.")

    tkMessageBox.showinfo("Success", "Success.\nThe output file was created successfully under {}.\n"
                                     "Quitting program.".format(outpath))
    exit(0)
    root.mainloop()



if __name__ == '__main__':
    main()