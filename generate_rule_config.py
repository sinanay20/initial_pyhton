#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: generate_rule_config
# Author: Samuel Bramm
# Date: 2018-04-13
# Description: Updates rule configuration with selectable servers for each tenant
# Version: v1.6
# v1.6 bugfix for missing parameters


from openpyxl import *
from openpyxl.styles import *
from ibm_db import exec_immediate
from Tkinter import *
from ttk import *
import ibm_db
import tkMessageBox
import yaml
import time
from tkFileDialog import askopenfilename
from tkSimpleDialog import askstring
import urllib2

# global variables
root = Tk()
config_path = './config/generate_rule_config.yml'  # TODO change assignments
tenant_list = []
tenant_list_old = []
server_data_list = []
default_tenant_server = ""
tenants_set = IntVar()
outfile_set = IntVar()
check_set = IntVar()
finish_set = IntVar()
default_server = []
output_path = ""
found_parameter_list = []
exclude_parameter_list = []
new_tenants = []


# TODO
# implement upload function
# make script ready for iCON scripts

def upload_rc():
    print "upload"
    global root
    ul_window = Toplevel(root)
    root.withdraw()
    # select old and new file
    # parse old and new file for tab names
    # display tabs which exist in both files (excluding first two tabs) with update button and server selection and tenant
    # iterate over selected tabs and and compare values, return list with [parameter, new_value, old_values]
    # iterate over update_list, build update commands and execute them on the specified server
    return None


def build_parameter_list():
    global found_parameter_list
    global output_path
    global exclude_parameter_list
    existing_parameters = []
    new_parameters = []
    try:
        book = load_workbook(output_path)
        sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
        for i in range(1, sheet.max_row + 1):
            param = sheet.cell(row=i, column=1).value
            if param is not None:
                if str(param.encode('utf-8', 'ignore')).lower().strip() not in exclude_parameter_list:
                    existing_parameters.append(str(param.encode('utf-8', 'ignore')).lower().strip())
    except IOError:
        tkMessageBox.showerror("IO/Error", "Outputfile %s not found" % (output_path))
    for found_parameter in found_parameter_list:
        if found_parameter.lower() not in existing_parameters:
            new_parameters.append(found_parameter)
    with open(os.path.dirname(output_path) + "/new_parameters" + time.strftime("_%Y_%m_%d") + ".txt",
              'a') as error_file:
        for new_parameter in new_parameters:
            error_file.write("%s\n" % new_parameter)
    if len(new_parameters) != 0:  # TODO verify if working correctly
        return True
    return None


def write_yaml():
    # write config to yaml file
    global config_path
    global server_data_list
    global new_tenants
    global tenant_list_old
    try:
        config = yaml.safe_load(open(config_path))
        # update version information
        for entry in server_data_list:
            config["servers"][entry[0]]["version"] = entry[4]
        # update new tenants
        if len(tenant_list_old) != len(tenant_list):
            for tenant in tenant_list:
                if tenant not in tenant_list_old:
                    config["tenants"].update({tenant: {'server': 'dev4'}})
        # write updates to file
        with open(config_path, 'w') as out_file:
            yaml.safe_dump(config, out_file, default_flow_style=False)
    except IOError:
        tkMessageBox.showerror("I/O Error", "The configuration file is not available under %s" % config_path)
    return None


def build_urllib_opener(proxy_url, proxy_user, proxy_pw):
    # build proxy opener for urllib2
    passwd_mgr = urllib2.HTTPPasswordMgrWithDefaultRealm()
    passwd_mgr.add_password(None, proxy_url, proxy_user, proxy_pw)
    proxy_handler = urllib2.ProxyHandler({'https': proxy_url})
    proxy_auth_handler = urllib2.ProxyBasicAuthHandler(passwd_mgr)
    urllib2.install_opener(urllib2.build_opener(proxy_handler, proxy_auth_handler))
    return None


def get_iCON_version(window):
    # retrieves iCON version from monitoring page
    version = ""
    global server_data_list
    global check_set
    server_list, user, pwd, port, version_list = zip(*server_data_list)
    version_list = list(version_list)
    # build proxy for selection
    proxy_user = askstring("Username", "Please specify your username (EMEA-ID):")
    proxy_pw = askstring("Username", "Please specify your password (Windows login):", show='*')
    build_urllib_opener('53.18.255.200:3128', proxy_user, proxy_pw) #53.71.122.84:3128
    # update version
    for i, server in enumerate(server_list):
        try:
            if server == "int":
                response = urllib2.urlopen(
                    'https://aftersales-int.i.daimler.com/icon-admin/monitoring/monitoring?monitoring=information')
            else:
                response = urllib2.urlopen(
                    'https://icon-dev03.emea.lab.corpintra.net/icon-admin/monitoring/monitoring?monitoring=information')
            html = response.read()
            version_pattern = re.compile(r'iCON-([0-9\.]*)')
            if re.search(version_pattern, html):
                found_version = re.search(version_pattern, html).group(1)
                version_list[i] = found_version
        except:
            tkMessageBox.showerror("Http Error",
                                   "Connection to server %s not possible. Using default value from configuration file." % server)
    server_data_list = [list(x) for x in zip(server_list, user, pwd, port, version_list)]
    check_set.set(1)
    return version


def write_excel_file(tenant, server, rule_config_result_set, md_version):
    # write results to excel file in new tab
    global output_path
    header = ['CODE', 'TENANTID', 'INSTANCEVERSION', 'PARAMETERNAME', 'VALIDFORTASKTYPE', 'PARAMETERVALUE',
              'FK_VALIDFORBRAND', 'FK_VALIDFORPRODUCT', 'FK_VALIDFORWORKSHOPINVOICEDEFINITION',
              'FK_VALIDFORCUSTOMERINVOICEDEFINITION', 'VALIDFORVEHICLECONTRACTSTATUS', 'VALIDFORDIVISION',
              'VALIDFORVEHICLECONTRACTVARIANTDEFINITION', 'VALIDFORFINANCIALDOCUMENTDEFINITION',
              'VALIDFORFINANCIALDOCUMENTSOURCE', 'VALIDFROM', 'VALIDTO', 'SELECTIONSORTKEY', 'ISDEFAULTSELECTION',
              'RELEASESTATUS', 'IGNORED', 'TECHCHANGETIME', 'TECHCHANGEUSER', 'TECHCREATETIME', 'TECHCREATEUSER']
    try:
        book = load_workbook(output_path)
        release_version = [x[4] for x in server_data_list if x[0] == server.lower()][0]
        sheet_name = "%s %s - %s" % (server.upper(), tenant.upper()[5:], release_version)
        book.create_sheet(sheet_name)
        working_sheet = book.get_sheet_by_name(sheet_name)
        working_sheet.append(header)
        # write rules in tab
        for rule in rule_config_result_set:
            working_sheet.append(rule)
        # update overview
        overview_sheet = book.get_sheet_by_name('Overview')
        custom_max_row = overview_sheet.max_row
        for i in range(1, overview_sheet.max_row + 2):
            if overview_sheet.cell(row=i, column=1).value == None:
                custom_max_row = i
                break
        overview_data = [sheet_name, server.upper(), tenant.upper()[5:], release_version, md_version,
                         time.strftime("%d.%m.%y")]
        for column in range(1, len(overview_data) + 1):
            overview_sheet.cell(row=custom_max_row, column=column).value = overview_data[column - 1]
            overview_sheet.cell(row=custom_max_row, column=column).border = Border(left=Side(style='thin'),
                                                                                   right=Side(style='thin'),
                                                                                   top=Side(style='thin'),
                                                                                   bottom=Side(style='thin'))
            overview_sheet.cell(row=custom_max_row, column=column).alignment = Alignment(horizontal="center",
                                                                                         vertical="center")
        book.save(output_path)
    except:
        tkMessageBox.showerror("IO/Error", "Outputfile %s not found" % (output_path))
    return None


def get_server_data(server):
    # return connection data for specified server
    global server_data_list
    user = ""
    pwd = ""
    port = ""
    for entry in server_data_list:
        if entry[0] == server:
            user = entry[1]
            pwd = entry[2]
            port = entry[3]
    return user, pwd, port


def download_config(update_list):
    # download rule configuration for each selected tenant
    global finish_set
    global found_parameter_list
    for entry in update_list:
        if entry[2].get() != "":
            rule_config_result_set = []
            tenant = entry[0]
            server = entry[1].get()
            user, pwd, port = get_server_data(server)
            result = tkMessageBox.askquestion("Updating rule configuration",
                                              "Updating rule configuration for tenant %s from server %s. Press \"yes\" if a connection via putty to the server is established." % (
                                                  tenant, server))
            if result == "yes":
                try:
                    connection = ibm_db.connect("DATABASE=icon;HOSTNAME=localhost;PORT=" + str(
                        port) + ";PROTOCOL=TCPIP;UID=" + user + ";PWD=" + pwd + ";", "", "")
                except:
                    tkMessageBox.showerror("Database Error",
                                           "The connection to " + server + " could not be established.\n Check the connection in putty and try again.")
                rule_config_query = "select * from icon.CONFIGURATION_RULECONFIGURATION where TENANTID='" + tenant + "'"
                rule_config_statement = exec_immediate(connection, rule_config_query)
                rule_config_results = ibm_db.fetch_tuple(rule_config_statement)
                while rule_config_results:
                    rule_config_result_set.append(
                        [str(rule_config_results[0]), str(rule_config_results[1]), str(rule_config_results[2]),
                         str(rule_config_results[3]), str(rule_config_results[4]), str(rule_config_results[5]),
                         str(rule_config_results[6]), str(rule_config_results[7]), str(rule_config_results[8]),
                         str(rule_config_results[0]), str(rule_config_results[9]), str(rule_config_results[10]),
                         str(rule_config_results[11]), str(rule_config_results[12]), str(rule_config_results[13]),
                         str(rule_config_results[14]), str(rule_config_results[15]), str(rule_config_results[16]),
                         str(rule_config_results[17]), str(rule_config_results[18]), str(rule_config_results[19]),
                         str(rule_config_results[20]), str(rule_config_results[21]), str(rule_config_results[22]),
                         str(rule_config_results[23]), str(rule_config_results[24])])
                    if str(rule_config_results[3]).strip() not in found_parameter_list:
                        found_parameter_list.append(str(rule_config_results[3]).strip())
                    rule_config_results = ibm_db.fetch_tuple(rule_config_statement)
                master_data_query = "SELECT * FROM icon.MASTERDATA_MASTERDATARELEASEINFO where TENANTID='" + tenant + "' ORDER BY ACTIVATIONDATE DESC FETCH FIRST 1 ROWS ONLY;"
                master_data_statement = exec_immediate(connection, master_data_query)
                master_data_results = ibm_db.fetch_tuple(master_data_statement)
                if master_data_results:
                    timestamp = str(master_data_results[5])[6:] + "." + str(master_data_results[5])[4:6] + "." + str(
                        master_data_results[5])[2:4]
                    master_data_version = "MD_" + tenant.upper()[5:] + "_" + str(
                        master_data_results[0]) + "/" + timestamp
                else:
                    master_data_version = ""
                write_excel_file(tenant, server, rule_config_result_set, master_data_version)
            else:
                pass
    finish_set.set(1)
    return None


def get_output_file():
    global output_path
    global outfile_set
    output_path = askopenfilename(initialdir=os.getcwd() + "/output/", title="Choose the Rule Configuration Excel file",
                                  filetypes=[("Excel", "*.xlsx")])
    outfile_set.set(1)
    return None


def display_tenant_selection(window, server_list):
    # display selection menu for each tenant
    global tenant_list
    global default_server
    t, s = zip(*default_server)
    # tenant header
    tenant_label = Label(window,
                         text="3. Select a server for each tenant you want to update and check the corresponding box.")
    tenant_label.grid(row=2, column=0, columnspan=3, padx=10, pady=10)
    Label(window, text="Tenant").grid(row=3, column=0, padx=10, pady=10)
    Label(window, text="Server").grid(row=3, column=1, padx=10, pady=10)
    Label(window, text="Update").grid(row=3, column=2, padx=10, pady=10)
    update_list = []
    button_list = []
    for idx, tenant in enumerate(tenant_list):
        tmp = StringVar(window)
        update = StringVar(window)
        Label(window, text=tenant).grid(row=(idx + 4), column=0, padx=10, pady=10)  # tenant
        if tenant in t:
            o = OptionMenu(window, tmp, s[t.index(tenant)], *sorted(server_list))
        else:
            o = OptionMenu(window, tmp, "dev4", *sorted(server_list))
        o.grid(row=(idx + 4), column=1, padx=10, pady=10)
        c = Checkbutton(window, variable=update)
        c.grid(row=(idx + 4), column=2, padx=10, pady=10)
        button_list.append([o, c])
        update_list.append([tenant, tmp, update])
    return update_list, button_list


def load_tenants(server_name):
    # load available tenants from server and update global tenant list
    global tenants_set
    global server_data_list
    global tenant_list
    for entry in server_data_list:
        if entry[0] == server_name:
            user = entry[1]
            pwd = entry[2]
            port = entry[3]
    try:
        connection = ibm_db.connect(
            "DATABASE=icon;HOSTNAME=localhost;PORT=" + str(port) + ";PROTOCOL=TCPIP;UID=" + user + ";PWD=" + pwd + ";",
            "", "")
    except:
        tkMessageBox.showerror("Database Error",
                               "The connection to " + server_name + " could not be established.\n Check the connection in putty and try again.")
    tenant_query = "select TENANTID from icon.CONFIGURATION_RULECONFIGURATION group by TENANTID"
    tenant_statement = exec_immediate(connection, tenant_query)
    tenant_result = ibm_db.fetch_tuple(tenant_statement)

    while tenant_result:
        if tenant_result[0] not in tenant_list:
            tenant_list.append(tenant_result[0])
        tenant_result = ibm_db.fetch_tuple(tenant_statement)
    tenants_set.set(1)
    return None


def download_rc():
    # Download Rule Configuration for selected tenants from the associated servers
    global root
    global server_data_list
    global default_tenant_server
    global tenants_set
    global tenant_list
    global outfile_set
    global check_set
    global new_tenants
    server_list, user, pwd, port, version = zip(*server_data_list)
    dl_window = Toplevel(root)
    root.withdraw()
    tenant_list_len_old = len(tenant_list)
    # display tenant update selection
    dl_window.title("Rule Configuration Downloader")
    tenant_label = Label(dl_window, text="1. Select server to update tenant list.")
    tenant_label.grid(row=0, column=0, pady=10, padx=10)
    tenant_server = StringVar(dl_window)

    tenant_dropdown = OptionMenu(dl_window, tenant_server, default_tenant_server, *sorted(server_list))
    tenant_dropdown.grid(row=0, column=1, pady=10, padx=10)
    tenant_button = Button(dl_window, text="load tenants", command=lambda: load_tenants(tenant_server.get()))
    tenant_button.grid(row=0, column=2, pady=10, padx=10)
    dl_window.wait_variable(tenants_set)
    tenant_button.config(state=DISABLED)

    if tenant_list_len_old != len(tenant_list):
        tkMessageBox.showinfo("New tenants detected",
                              "There are %d tenants available on the server %s, but only %d tenants are in the config file. Please consider updating the file" % (
                                  len(tenant_list), tenant_server.get(), tenant_list_len_old))
    # update version for dev server
    check_label = Label(dl_window, text="2. Check for the release versions for each tenant.")
    check_label.grid(row=1, column=0, columnspan=2, pady=10, padx=10)
    check_button = Button(dl_window, text="check", command=lambda: get_iCON_version(dl_window))
    check_button.grid(row=1, column=2, pady=10, padx=10)
    dl_window.wait_variable(check_set)
    check_button.config(state=DISABLED)
    # display tenant and server selection
    update_list, button_list = display_tenant_selection(dl_window, server_list)
    update_label = Label(dl_window, text="4. Select a output file.")
    update_label.grid(row=len(tenant_list) + 4, column=0, columnspan=2, padx=10, pady=10)
    update_button = Button(dl_window, text="select", command=lambda: get_output_file())
    update_button.grid(row=len(tenant_list) + 4, column=2, padx=10, pady=10)
    dl_window.wait_variable(outfile_set)
    # disable previous selections
    update_button.config(state=DISABLED)
    [x[1].config(state=DISABLED) for x in button_list]
    [x[0].config(state=DISABLED) for x in button_list]
    # display download button
    downlad_label = Label(dl_window, text="5. Press the download button and follow the instructions.")
    downlad_label.grid(row=len(tenant_list) + 5, column=0, columnspan=2, padx=10, pady=10)
    download_button = Button(dl_window, text="download", command=lambda: download_config(update_list))
    download_button.grid(row=len(tenant_list) + 5, column=2, padx=10, pady=10)
    dl_window.wait_variable(finish_set)
    download_button.config(state=DISABLED)
    write_yaml()
    ret = build_parameter_list()
    if ret:
        if tkMessageBox.askyesno("Sucess",
                                 "All tenants have been updated and safed to the configuration file. There were new parameters which have been saved to a file along with the configuration file. Do you want to quit?"):
            exit(0)
    else:
        if tkMessageBox.askyesno("Sucess",
                                 "All tenants have been updated and safed to the configuration file. Do you want to quit?"):
            exit(0)
    return None


def get_config():
    # Load config from config file
    global config_path
    global tenant_list
    global server_data_list
    global default_tenant_server
    global default_server
    global exclude_parameter_list
    global tenant_list_old
    try:
        config = yaml.safe_load(open(config_path))
        for tenant in config["tenants"]:
            tenant_list.append(tenant)
            tenant_list_old.append(tenant)
            default_server.append([tenant, config["tenants"][tenant]["server"]])
        for server in config["servers"]:
            server_data_list.append([server, config["servers"][server]["user"], config["servers"][server]["pwd"],
                                     config["servers"][server]["port"], config["servers"][server]["version"]])
        default_tenant_server = config["general"]["default_tenant_server"]
        exclude_parameter_list = config["general"]["exclude_parameter_list"]
    except IOError:
        tkMessageBox.showerror("I/O Error", "The configuration file is not available under %s" % config_path)
    return None


def main():
    global root
    # load values (tenants, server connection) from config file
    get_config()
    # load screen for tenant selection
    root.title("Rule Configuration Helper")
    selection = Label(root, text="Choose what you want to do!")
    selection.grid(row=0, column=0, pady=10, padx=10, columnspan=2)
    download_label = Label(root, text="Download Rule Configuration from servers.")
    # upload_label = Label(root, text="Upload Rule Configuration to servers.")
    upload_label = Label(root, text="Upload Rule Configuration to servers.\n Not yet implemented.")  # TODO revert
    download_label.grid(row=1, column=0, pady=10, padx=10)
    upload_label.grid(row=1, column=1, pady=10, padx=10)
    upload_button = Button(root, text="upload", command=lambda: upload_rc())
    download_button = Button(root, text="download", command=lambda: download_rc())
    upload_button.grid(row=2, column=1, pady=10, padx=10)
    download_button.grid(row=2, column=0, pady=10, padx=10)
    upload_button.config(state=DISABLED)  # TODO revert
    root.mainloop()
    exit(0)


if __name__ == '__main__':
    main()
