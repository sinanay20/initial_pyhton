#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: Salesman xlsx to xml
# Author: Benjamin Schmelcher
# Date: 10.09.2018
# Description: Provides a salesman.xml file based on a error file (format see documentation)
# Version: 0.0.1

import csv
import os
import tkMessageBox
from Tkinter import IntVar, Tk
from tkFileDialog import askopenfilename
from ttk import Label, Button

from openpyxl import load_workbook

#output_file_name = "./output/Salesman.xml"
#xlsx_file_name = "./Copy of errors_2018_09_06 (002).xlsx"
salesmen = []

root = Tk()
file_selected = IntVar()
xlsx_file_name = ""
output_file_name = ""


def create_xml_file():
    global output_file_name

    try:
        output_writer = open(output_file_name, 'w')
    except:
        tkMessageBox.showerror("I/O Error",
                               "Can not open file" + os.path.dirname(output_file_name) + "/" + output_file_name + "for writing")

        exit(1)


    #create Header in file
    output_writer.write(
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n"
        "<common:ServiceInvocationCollection xmlns:common=\"http://common.icon.daimler.com/il\""
        "xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" "
        "xmlns:partner_pl=\"http://partner.icon.daimler.com/pl\" "
        "xmlns:mdsd_sl=\"http://system.mdsd.ibm.com/sl\" "
        "xmlns:xsd=\"http://www.w3.org/2001/XMLSchema\">\n"
        "  <!--Related to CIM: 20140811_CIM_EDF_PhysicalPerson(salesman)_Mig_BEL_WavePreInt4_iter1_v1.0.xlsx-->\n"
        "  <!--Related to Masterdata:5.11-->\n"
        "  <!--Source-database: db3.S415VM779.tst-->\n"
        "  <executionSettings xsi:type=\"mdsd_sl:ExecutionSettingsType\" "
        "dateTime=\"2018-09-06T00:00:00\" userId=\"ICON_tec_EDF01\" "
        "tenantId=\"81930CN\" causation=\"migration\" "
        "additionalInformation1=\"1\" correlationId=\"initial export\" "
        "issueThreshold=\"error\"/>\n"
        )

    for salesman in salesmen:
        #physical-person
        output_writer.write(
        "  <invocation operation=\"createPhysicalPerson\">\n"
        "    <parameter xsi:type=\"partner_pl:PhysicalPersonType\" "
        "externalId=\""+salesman[0]+"\" "
        "sourceSystem=\"migration\" "
        "masterDataReleaseVersion=\"9\" "
        "partnerType=\"salesman\" "
        "migrationDate=\"2018-09-06T00:00:00\" "
        "firstName=\""+salesman[1]+"\" "
        "lastName=\""+salesman[2]+"\" "
        "isUserLastLogin=\"false\" "
        "dealerDirectoryUid=\""+salesman[0]+"\" "
        "organisationalLevel=\"mpc\">\n"
        "      <communicationData"
        )
        if(salesman[3] != None):
            output_writer.write(
            " phoneNumber=\""+salesman[3]+"\""
            )
        if(salesman[4] != None):
            output_writer.write(
            " mobile=\"016022334455\""
            )
        if(salesman[5] != None):
            output_writer.write(
            " email=\"test.test@daimler.com\""
            )
        output_writer.write(
            "/>\n    </parameter>\n"
            "  </invocation>\n"
            )

    output_writer.write(
    "</common:ServiceInvocationCollection >"
    )




def read_xlsx_file():
    global xlsx_file_name
    global salesmen
    book = load_workbook(xlsx_file_name)
    sheets = book.get_sheet_names()
    sheet = book.get_sheet_by_name("user data")

    for index, row in enumerate(sheet.iter_rows()):
        tempsalesman = []
        if index>0:
            for cell in row:
                tempsalesman.append(cell.value)
            salesmen.append(tempsalesman)


def select_error_file():
    global root
    global xlsx_file_name
    global output_file_name
    xlsx_file_name = askopenfilename(parent=root, title="Select error file", filetypes=[("Excel", "*.xlsx")])
    output_file_name = os.path.dirname(xlsx_file_name)+"/Salesman.xml"

    file_selected.set(1)


def main():
    global root
    global file_selected
    global output_file_name
    selectLabel = Label(root, text="1. Select the errorfile containing the information to create the Salesman xml file.")
    selectLabel.grid(row=0, column=0, padx=10, pady=10)
    selectButton = Button(root, text="select", command=lambda: select_error_file())
    selectButton.grid(row=0, column=1, padx=10, pady=10)

    root.wait_variable(file_selected)


    read_xlsx_file()
    create_xml_file()

    tkMessageBox.showinfo(title="Success", message="Succes. The outputfile has been created in {}.".format(output_file_name))

    exit(0)





if __name__ == '__main__':
    main()