#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: analyze_I23
# Author: Samuel Bramm
# Date: 2018-01-12
# Description: Analyze SSLART and SSLTEIL
# Version: v1.0

from Tkinter import *
from ttk import *
from tkFileDialog import askopenfilename
import os
import tkMessageBox
import csv
import sys
import time
from openpyxl import *
from openpyxl.styles import *

root_window = Tk()
sslart_path = ""
sslteil_path = ""
db_damage_codes_pre_import_path = ""
db_damage_codes_after_import_path = ""
db_damage_types_pre_import_path = ""
db_damage_types_after_import_path = ""

def write_results(len_damage_codes_from_file,len_damage_codes_pre_import, damage_code_results, len_damage_types_from_file, len_damage_types_pre_import, damage_type_results, sub_window, double_damage_types):
    book = Workbook()
    global root_window
    out_file = os.getcwd() + "/output/I23_analysis_" + time.strftime("%Y_%m_%d_") + ".xlsx"
    sheet = book.active
    sheet.title = "I23 Analysis"
    #Damagecodes
    sheet['A1'] = "Damagecodes"
    sheet['A1'].fill = PatternFill(fgColor='52BE80', fill_type="solid")
    sheet['B1'].fill = PatternFill(fgColor='52BE80', fill_type="solid")
    sheet.merge_cells('A1:E1')
    sheet['A2'] = "in new file"
    sheet['B2'] = len_damage_codes_from_file
    sheet['A3'] = "known in DEV"
    sheet['B3'] = len_damage_codes_pre_import
    sheet['A5'] = "difference"
    sheet['B5'] = abs(len_damage_codes_from_file-len_damage_codes_pre_import)
    sheet['A6'] = "new damagecodes"
    sheet['B6'] = len(damage_code_results[0])
    sheet['A7'] = "deleted damagecodes"
    sheet['B7'] = len(damage_code_results[1])
    sheet['A10'] = "instanceversions pre import"
    sheet['A10'].fill = PatternFill(fgColor="F9E79F", fill_type="solid")
    sheet['B10'].fill = PatternFill(fgColor="F9E79F", fill_type="solid")
    sheet.merge_cells('A10:B10')
    sheet['A11'] = "instanceversion"
    sheet['B11'] = "count"
    index = 12
    for i,val in enumerate(damage_code_results[2]):
        sheet.cell(row=index, column=1).value = i
        sheet.cell(row=index, column=2).value = damage_code_results[2][i]
        index += 1
    sheet['D10'] = "instanceversions after import"
    sheet['D10'].fill = PatternFill(fgColor="F9E79F", fill_type="solid")
    sheet['E10'].fill = PatternFill(fgColor="F9E79F", fill_type="solid")
    sheet.merge_cells('D10:E10')
    sheet['D11'] = "instanceversion"
    sheet['E11'] = "count"
    index = 12
    for j,val in enumerate(damage_code_results[3]):
        sheet.cell(row=index, column=4).value = j
        sheet.cell(row=index, column=5).value = damage_code_results[3][j]
        index += 1
    #Damagetypes
    sheet['G1'] = "Damagetypes"
    sheet['G1'].fill = PatternFill(fgColor='52BE80', fill_type="solid")
    sheet['H1'].fill = PatternFill(fgColor='52BE80', fill_type="solid")
    sheet.merge_cells('G1:K1')
    sheet['G2'] = "in new file"
    sheet['H2'] = len_damage_types_from_file
    sheet['G3'] = "known in DEV"
    sheet['H3'] = len_damage_types_pre_import
    sheet['G5'] = "difference"
    sheet['H5'] = abs(len_damage_types_from_file-len_damage_types_pre_import)
    sheet['G6'] = "new damagetypes"
    sheet['H6'] = len(damage_type_results[0])
    sheet['G7'] = "deleted damagetypes"
    sheet['H7'] = len(damage_type_results[1])
    sheet['G8'] = "double damagetypes in file"
    sheet['H8'] = double_damage_types
    sheet['G10'] = "instanceversions pre import"
    sheet['G10'].fill = PatternFill(fgColor="F9E79F", fill_type="solid")
    sheet['H10'].fill = PatternFill(fgColor="F9E79F", fill_type="solid")
    sheet.merge_cells('G10:H10')
    sheet['G11'] = "instanceversion"
    sheet['H11'] = "count"
    index = 12
    for i, val in enumerate(damage_type_results[2]):
        sheet.cell(row=index, column=7).value = i
        sheet.cell(row=index, column=8).value = damage_type_results[2][i]
        index += 1
    sheet['J10'] = "instanceversions after import"
    sheet['J10'].fill = PatternFill(fgColor="F9E79F", fill_type="solid")
    sheet['K10'].fill = PatternFill(fgColor="F9E79F", fill_type="solid")
    sheet.merge_cells('J10:K10')
    sheet['J11'] = "instanceversion"
    sheet['K11'] = "count"
    index = 12
    for i, val in enumerate(damage_type_results[3]):
        sheet.cell(row=index, column=10).value = i
        sheet.cell(row=index, column=11).value = damage_type_results[3][i]
        index += 1
    sheet2 = book.create_sheet(title="new and deleted damagecode")
    sheet2['A1'] = "new codes"
    for i,new in enumerate(damage_code_results[0]):
        sheet2.cell(row=(i+2), column=1).value = new
    sheet2['C1'] = "deleted codes"
    for i,deleted in enumerate(damage_code_results[1]):
        sheet2.cell(row=(i + 2), column=3).value = deleted
    sheet3 = book.create_sheet(title="new and deleted damagetypes")
    sheet3['A1'] = "new types"
    for i, new in enumerate(damage_type_results[0]):
        sheet3.cell(row=(i + 2), column=1).value = new
        sheet3['C1'] = "deleted types"
    for i, deleted in enumerate(damage_type_results[1]):
        sheet3.cell(row=(i + 2), column=3).value = deleted
    book.save(out_file)
    tkMessageBox.showinfo("Sucess","All files parsed. Output file created in " + out_file)
    sub_window.quit()
    root_window.quit()
    return None


def compare_damage_types(damage_types_from_file, damage_types_pre_import, damage_types_after_import, sub_window):
    new_values = []
    deleted_values = []
    pre_values, pre_instance = zip(*damage_types_pre_import)
    pre_instance = [int(x) for x in pre_instance]
    instance_count_pre_import = [0] * (max(pre_instance) + 1)
    for damage_type in damage_types_from_file:
        sub_window.update()
        if damage_type not in pre_values:
            new_values.append(damage_type)
    after_values, after_instance = zip(*damage_types_after_import)
    after_instance = [int(x) for x in after_instance]
    instance_count_after_import = [0] * (max(after_instance) + 1)
    for deleted_type in pre_values:
        sub_window.update()
        if deleted_type not in damage_types_from_file:
            deleted_values.append(deleted_type)
    for p_instance in pre_instance:
        instance_count_pre_import[p_instance] += 1
    for a_instance in after_instance:
        instance_count_after_import[a_instance] += 1
    return [new_values, deleted_values, instance_count_pre_import, instance_count_after_import]

def read_damage_types_after_import(sub_window):
    global db_damage_types_after_import_path
    damage_types = []
    with open(db_damage_types_after_import_path, 'rb') as types_file:
        types_reader = csv.reader(types_file, delimiter=";")
        for i,line in enumerate(types_reader):
            sub_window.update()
            if i > 0 and [line[0], line[1]] not in damage_types:
                damage_types.append([line[0],line[1]])
    return damage_types

def read_damage_types_pre_import(sub_window):
    damage_types = []
    global db_damage_types_pre_import_path
    with open(db_damage_types_pre_import_path, 'rb') as types_file:
        types_reader = csv.reader(types_file, delimiter=";")
        for i,line in enumerate(types_reader):
            sub_window.update()
            if i > 0 and [line[0],line[1]] not in damage_types:
                damage_types.append([line[0],line[1]])
    return damage_types

def read_sslart(sub_window):
    global sslart_path
    damage_types = []
    duplicates = 0
    uniq = []
    with open(sslart_path,'rb') as types_file:
        types_reader = csv.reader(types_file, delimiter=" ")
        for line in types_reader:
            sub_window.update()
            if line[0] == '01':
                if line[1] == 'M' or line[1] == 'L' or line[1] == 'X':
                    damage_types.append(line[2])
                else:
                    damage_types.append(line[3])
    for posible_double in damage_types:
        if posible_double in uniq:
            duplicates += 1
        else:
            uniq.append(posible_double)
    return damage_types, duplicates

def compare_damage_codes(damage_codes_from_file, damage_codes_pre_import, damage_codes_after_import, sub_window):
    new_values = []
    deleted_values = []
    pre_values, pre_instance = zip(*damage_codes_pre_import)
    pre_instance = [int(x) for x in pre_instance]
    instance_count_pre_import = [0] * (max(pre_instance) + 1)
    for damage_code in damage_codes_from_file:
        sub_window.update()
        if damage_code not in pre_values:
            new_values.append(damage_code)
    after_values, after_instance = zip(*damage_codes_after_import)
    after_instance = [int(x) for x in after_instance]
    instance_count_after_import = [0] * (max(after_instance) + 1)
    for deleted_code in pre_values:
        sub_window.update()
        if deleted_code not in damage_codes_from_file:
            deleted_values.append(deleted_code)
    for p_instance in pre_instance:
        instance_count_pre_import[p_instance] += 1
    for a_instance in after_instance:
        instance_count_after_import[a_instance] += 1
    return [new_values, deleted_values, instance_count_pre_import, instance_count_after_import]

def read_damage_codes_after_import(sub_window):
    global db_damage_codes_after_import_path
    damage_codes = []
    with open(db_damage_codes_after_import_path, 'rb') as db_file:
        damage_reader = csv.reader(db_file,delimiter=";")
        for i,line in enumerate(damage_reader):
            sub_window.update()
            if i > 0:
                damage_codes.append([line[0], line[1]])
    return damage_codes

def read_damage_codes_pre_import(sub_window):
    damage_codes = []
    global db_damage_codes_pre_import_path
    with open(db_damage_codes_pre_import_path, 'rb') as db_file:
        damage_reader = csv.reader(db_file, delimiter=";")
        for i, line in enumerate(damage_reader):
            sub_window.update()
            if i > 0:
                damage_codes.append([line[0],line[1]])
    return damage_codes

def read_sslteil(sub_window):
    global sslteil_path
    damage_codes = []
    csv.field_size_limit(sys.maxsize)
    with open(sslteil_path, 'rb') as damage_file:
        damage_reader = csv.reader(damage_file, delimiter=" ")
        for line in damage_reader:
            sub_window.update()
            if line[1] != '' and line[1] not in damage_codes:
                damage_codes.append(line[1])
    return damage_codes

def read_and_analyze():
    sub_window = Toplevel()
    sub_window.title("Analyzing")
    Label(sub_window,text="1. Analyzing damagecodes. Please be patient.").grid(row=0,column=0,padx=10,pady=10)
    p = Progressbar(sub_window, orient=HORIZONTAL, length=100, mode='indeterminate')
    p.grid(row=0, column=1, padx=10, pady=10)
    p.start()
    sub_window.update()
    # lese damagecodes aus Datei
    damage_codes_from_file = read_sslteil(sub_window)
    # lese damagecodes aus db (pre import)
    damage_codes_pre_import = read_damage_codes_pre_import(sub_window)
    # lese damagecodes aus db (after import)
    damage_codes_after_import = read_damage_codes_after_import(sub_window)
    # vergleiche welche damagecode neu hinzu gekommen sind
    damage_code_results = compare_damage_codes(damage_codes_from_file, damage_codes_pre_import, damage_codes_after_import, sub_window)
    p.stop()
    Label(sub_window,text="2. Analyzing damagetypes. Please be patient.").grid(row=1,column=0,padx=10,pady=10)
    p1 = Progressbar(sub_window, orient=HORIZONTAL, length=100, mode='indeterminate')
    p1.grid(row=1, column=1, padx=10, pady=10)
    p1.start()
    sub_window.update()
    # lese damagetypes aus datei
    damage_types_from_file, double_damage_types = read_sslart(sub_window)
    # lese damgetypes aus db (pre import)
    damage_types_pre_import = read_damage_types_pre_import(sub_window)
    # lese damgetypes aus db (after import)
    damage_types_after_import = read_damage_types_after_import(sub_window)
    # vergleiche damagetypes
    damage_type_results = compare_damage_types(damage_types_from_file, damage_types_pre_import, damage_types_after_import, sub_window)
    p1.stop()
    Label(sub_window, text="3. Writing results to file.").grid(row=2, column=0, padx=10, pady=10)
    sub_window.update()
    write_results(len(damage_codes_from_file),len(damage_codes_pre_import), damage_code_results, len(damage_types_from_file), len(damage_types_pre_import), damage_type_results, sub_window, double_damage_types)
    return None

def check_files():
    global sslteil_path
    global sslart_path
    global db_damage_codes_after_import_path
    global db_damage_codes_pre_import_path
    global db_damage_types_after_import_path
    global db_damage_types_pre_import_path
    if sslart_path == '':
        tkMessageBox.showerror("I/O Error", "Please specify the location for the SSLART file.")
    if sslteil_path == '':
        tkMessageBox.showerror("I/O Error", "Please specify the location for the SSLTEIL file.")
    if db_damage_codes_pre_import_path == '':
        tkMessageBox.showerror("I/O Error", "Please specify the location for the damagecode (pre import) database file.")
    if db_damage_codes_after_import_path == '':
        tkMessageBox.showerror("I/O Error", "Please specify the location for the damagecode (after import) database file.")
    if db_damage_types_pre_import_path == '':
        tkMessageBox.showerror("I/O Error", "Please specify the location for the damagetypoe (pre import) database file.")
    if db_damage_types_after_import_path == '':
        tkMessageBox.showerror("I/O Error", "Please specify the location for the damagetypoe (after import) database file.")
    if os.path.isfile(sslteil_path)  and os.path.isfile(sslart_path) and os.path.isfile(db_damage_codes_pre_import_path) and os.path.isfile(db_damage_codes_after_import_path) and os.path.isfile(db_damage_types_pre_import_path) and os.path.isfile(db_damage_types_after_import_path):
        read_and_analyze()

    return None

def select_db_damage_type_after_import():
    global db_damage_types_after_import_path
    db_damage_types_after_import_path = askopenfilename(title="Select the database file", filetypes=[("CSV", "*.csv")])
    return None

def select_db_damage_type_pre_import():
    global db_damage_types_pre_import_path
    db_damage_types_pre_import_path = askopenfilename(title="Select the database file", filetypes=[("CSV", "*.csv")])
    return None

def select_db_damage_code_after_import():
    global db_damage_codes_after_import_path
    db_damage_codes_after_import_path = askopenfilename(title="Select the database file", filetypes=[("CSV", "*.csv")])
    return None

def select_db_damage_code_pre_import():
    global db_damage_codes_pre_import_path
    db_damage_codes_pre_import_path = askopenfilename(title="Select the database file", filetypes=[("CSV", "*.csv")])
    return None

def select_sslteil():
    global sslteil_path
    sslteil_path = askopenfilename(title="Select the SSLTEIL file",filetypes=[("SSLTEIL", "SSLTEIL.txt")])
    return None

def select_sslart():
    global sslart_path
    sslart_path = askopenfilename(title="Select the SSLART file",filetypes=[("SSLART", "SSLART.txt")])
    return None

def main():
    print sys.maxsize
    global root_window
    root_window.title("I23 Analyzer")
    Label(root_window,text="1. Select SSALRT file.").grid(row=0,column=0,padx=10,pady=10)
    Button(root_window,text="Open SSLART",command=select_sslart).grid(row=0,column=1,padx=10,pady=10)
    Label(root_window, text="2. Select SSLTEIL file.").grid(row=1, column=0, padx=10, pady=10)
    Button(root_window, text="Open SSLTEIL", command=select_sslteil).grid(row=1, column=1, padx=10, pady=10)
    Label(root_window, text="3. Select damagecode database export before import of files.").grid(row=2, column=0, padx=10, pady=10)
    Button(root_window, text="Open database", command=select_db_damage_code_pre_import).grid(row=2, column=1, padx=10, pady=10)
    Label(root_window, text="4. Select damagecode database export after importing the files.").grid(row=3, column=0, padx=10, pady=10)
    Button(root_window, text="Open database", command=select_db_damage_code_after_import).grid(row=3, column=1, padx=10, pady=10)
    Label(root_window, text="5. Select damagetype database export before import of files.").grid(row=4, column=0,padx=10, pady=10)
    Button(root_window, text="Open database", command=select_db_damage_type_pre_import).grid(row=4, column=1, padx=10, pady=10)
    Label(root_window, text="6. Select damagetype database export after importing the files.").grid(row=5, column=0,padx=10, pady=10)
    Button(root_window, text="Open database", command=select_db_damage_type_after_import).grid(row=5, column=1, padx=10, pady=10)
    Label(root_window, text="4. Press the analyze button.").grid(row=6, column=0, padx=10,pady=10)
    Button(root_window, text="analyze", command=check_files).grid(row=6, column=1, padx=10, pady=10)
    root_window.mainloop()
if __name__ == '__main__':
    main()