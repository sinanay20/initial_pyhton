#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: check_salesman.py
# Author: Samuel Bramm
# Co-Author: Benjamin Schmelcher
# Date: 2018-06-22
# Description: creates EDF xml files based on aat export and current users in database
# Version: v1.0


from Tkinter import *
from ttk import *
from tkFileDialog import askopenfilename
import tkMessageBox
import yaml
import time
import csv
from openpyxl import *

from patterns.db_connection import db_connection

root = Tk()
input_file_path = ""
file_set = IntVar()
load_done = IntVar()
found_users_db = []
tenants = []
connection_established = IntVar()
config_file_path = "./config/check_salesman.yml"
connection = db_connection()


def write_error_files(user_errors, org_errors):
    # write errors to excel file
    global input_file_path
    out_file_path = os.path.dirname(input_file_path) + "/errors_" + time.strftime("%Y_%m_%d") + ".xlsx"
    # TODO write errors to excel file
    if len(user_errors) != 0 or len(org_errors) != 0:
        book = Workbook()
        sheet = book.active
        sheet.title = "Salesman - Dealer Assignment"
        sheet.append(['UID', 'first name', 'name', 'phone', 'mobile', 'mail', 'homeorg'])
        for user in user_errors:
            sheet.append(user)

        sheet = book.create_sheet(title="Dealer data")
        sheet.append(
            ['GSSNOUTLETID', 'EXTERNALID', 'COMPANYNAME', 'COMPANYNAME2', 'VATID', 'GSSNCOMPANYOUTLETID', 'CITY',
             'COUNTRY', 'STREET', 'ZIPCODE', 'PHONE', 'FAX'])
        for idx, org in enumerate(org_errors):
            sheet.cell(row=idx + 2, column=1).value = org[0]
            for y, user in enumerate(org[1]):
                sheet.cell(row=idx + 2, column=13)
        sheet = book.create_sheet(title="user data")
        sheet.append(['UID', 'first name', 'name', 'phone', 'mobile', 'mail', 'homeorg'])
        for org in org_errors:
            for user in org[1]:
                sheet.append(user)
        try:
            book.save(out_file_path)
        except IOError:
            tkMessageBox.showerror("I/O Error", "Error writing to path {}".format(out_file_path))
        return out_file_path
    else:
        return False


def get_org_data(tenant, org):
    # connect to database and fetch data for given gssn

    data_stmt = "SELECT pp.EXTERNALID, pp.COMPANYNAME, pp.COMPANYNAME2, pp.VATID, pp.GSSNOUTLETCOMPANYID, pp.GSSNOUTLETOUTLETID, pa.CITY, pa.COUNTRY, pa.STREET, pa.ZIPCODE, pc.PHONENUMBER, pc.FAXNUMBER from icon.PARTNER_PARTNER pp JOIN icon.PARTNER_ADDRESS pa ON pp.LEGALADDRESS_OBJECTID = pa.OBJECTID JOIN icon.PARTNER_COMMUNICATIONDATA pc on pp.COMMUNICATIONDATA_OBJECTID = pc.OBJECTID WHERE pp.OBJECTINSTANCETYPE = 'OrganisationalPerson' AND pp.GSSNOUTLETOUTLETID = \'" + org + "\' AND pp.TENANTID = \'" + tenant + "\' AND pa.TENANTID = \'" + tenant + "\' AND pc.TENANTID = \'" + tenant + "\';"

    data_results = connection.execute_query(data_stmt)

    if data_results:
        clean_results = []
        for x in list(data_results):
            if x is None:
                clean_results.append("")
            else:
                tmplist = []
                for entry in x:
                    if entry is not None:
                        clean_results.append(entry.encode('utf-8'))
                    else:
                        clean_results.append("")

        # data_results = [x.encode('utf-8') for x in data_results if x is not None]
        return clean_results
    else:
        return data_results


def create_dealer_xml_files(user_data, tenant):
    # create dealer xml files if possible
    global input_file_path
    today = time.strftime("%Y-%m-%d")
    out_file_path = os.path.dirname(
        input_file_path) + "/02-OrganisationalPerson_" + today + "T000000_" + tenant.upper() + "_ICON_tec_EDF01_Dealer-00001.xml"
    user_errors = []
    org_errors = []
    uniqe_orgs = []
    # sort orgs
    for u in user_data:
        if u[6] == '':
            user_errors.append(u)
        elif u[6] not in uniqe_orgs:
            uniqe_orgs.append(u[6])

    with open(out_file_path, 'a') as dealer_file:
        # write header
        dealer_file.write("<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n")
        dealer_file.write(
            "<common:ServiceInvocationCollection xmlns:common=\"http://common.icon.daimler.com/il\" xmlns:partner_pl=\"http://partner.icon.daimler.com/pl\" xmlns:mdsd_sl=\"http://system.mdsd.ibm.com/sl\" xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xmlns:xsd=\"http://www.w3.org/2001/XMLSchema\">\n")
        dealer_file.write(
            "  <!--Related to CIM: 20140811_CIM_EDF_OrganisationalPerson(dealer)_Mig_BEL_WavePreInt4_iter1_v1.0.xlsx-->\n")
        dealer_file.write("  <!--Related to Masterdata:5.11-->\n")
        dealer_file.write("  <!--Source-database: db3.S415VM779.tst-->\n")
        dealer_file.write(
            "  <executionSettings xsi:type=\"mdsd_sl:ExecutionSettingsType\" dateTime=\"" + today + "T00:00:00\" userId=\"ICON_tec_EDF01\" tenantId=\"" + tenant + "\" causation=\"migration\" additionalInformation1=\"1\" issueThreshold=\"error\"/>\n")

        for idx, org in enumerate(uniqe_orgs):
            org_data = get_org_data(tenant, org)
            if org_data:
                dealer_file.write("  <invocation operation=\"updateOrganisationalPerson\">\n")
                dealer_file.write(
                    "    <parameter xsi:type=\"partner_pl:OrganisationalPersonType\" externalId=\"" + org_data[
                        0] + "\" sourceSystem=\"migration\" masterDataReleaseVersion=\"9\" partnerType=\"dealer\" state=\"active\" migrationDate=\"" + today + "T00:00:00\" companyName=\"" +
                    org_data[1] + "\" companyName2=\"" + org_data[2] + "\" companyInternal=\"no\" vatId=\"" + org_data[
                        3] + "\" gssnOutletCompanyId=\"" + org_data[4] + "\" gssnOutletOutletId=\"" + org + "\">\n")
                dealer_file.write(
                    "      <communicationData phoneNumber=\"" + org_data[10] + "\" faxNumber=\"" + org_data[
                        11] + "\"/>\n")
                dealer_file.write(
                    "      <legalAddress city=\"" + org_data[6] + "\" country=\"" + org_data[7] + "\" street=\"" +
                    org_data[8] + "\" zipCode=\"" + org_data[9] + "\"/>\n")
                for user in user_data:
                    if user[6] == org:
                        dealer_file.write("        <contactPartnerAssignment internal=\"false\" salesman=\"true\">\n")
                        dealer_file.write(
                            "          <contactPerson xsi:type=\"partner_pl:PhysicalPersonType\" externalId=\"" + user[
                                0].upper() + "\" sourceSystem=\"migration\"/>\n")
                        dealer_file.write("        </contactPartnerAssignment>\n")
                dealer_file.write("    </parameter>\n")
                dealer_file.write("  </invocation>\n")
            else:
                org_users = [u for u in user_data if u[6] == org]
                org_errors.append([org, org_users])
        # finalize file
        dealer_file.write("</common:ServiceInvocationCollection>")
    return out_file_path, user_errors, org_errors


def query_user_data(users):
    # query additional data for users from file
    global input_file_path
    user_data = []
    with open(input_file_path, 'rb') as input_file:
        input_reader = csv.reader(input_file, delimiter=";")
        for line in input_reader:
            if line[0] in users:
                phone = str(line[20])
                mobile = str(line[19])
                if 'E' in phone:
                    phone = ''
                if 'E' in mobile:
                    mobile = ''
                # [uid, first_name, name, phone, mobile, email, homeorg]
                user_data.append([line[0], line[1], line[2], phone, mobile, line[18], line[9]])
    return user_data


def create_salesman_xml_files(new_users, tenant):
    # create xml files for dealer and salesman
    global input_file_path
    today = time.strftime("%Y-%m-%d")
    out_file_path = os.path.dirname(
        input_file_path) + "/01-PhysicalPerson_" + today + "T000000_" + tenant.upper() + "_ICON_tec_EDF01_Salesman_00001.xml"
    with open(out_file_path, 'a') as salesman_output:
        # write header
        salesman_output.write("<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n")
        salesman_output.write(
            "<common:ServiceInvocationCollection xmlns:common=\"http://common.icon.daimler.com/il\" xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" xmlns:partner_pl=\"http://partner.icon.daimler.com/pl\" xmlns:mdsd_sl=\"http://system.mdsd.ibm.com/sl\" xmlns:xsd=\"http://www.w3.org/2001/XMLSchema\">\n")
        salesman_output.write(
            "  <!--Related to CIM: 20140811_CIM_EDF_PhysicalPerson(salesman)_Mig_BEL_WavePreInt4_iter1_v1.0.xlsx-->\n")
        salesman_output.write("  <!--Related to Masterdata:5.11-->\n")
        salesman_output.write("  <!--Source-database: db3.S415VM779.tst-->\n")
        salesman_output.write(
            "  <executionSettings xsi:type=\"mdsd_sl:ExecutionSettingsType\" dateTime=\"" + today + "T00:00:00\" userId=\"ICON_tec_EDF01\" tenantId=\"" + tenant + "\" causation=\"migration\" additionalInformation1=\"1\" correlationId=\"initial export\" issueThreshold=\"error\"/>\n")

        # write salesman
        for user in new_users:
            salesman_output.write("  <invocation operation=\"createPhysicalPerson\">\n")
            salesman_output.write("    <parameter xsi:type=\"partner_pl:PhysicalPersonType\" externalId=\"" + user[
                0].upper() + "\" sourceSystem=\"migration\" masterDataReleaseVersion=\"9\" partnerType=\"salesman\" migrationDate=\"" + today + "T00:00:00\" firstName=\"" +
                                  user[1] + "\" lastName=\"" + user[
                                      2] + "\" isUserLastLogin=\"false\" dealerDirectoryUid=\"" + user[
                                      0].upper() + "\" organisationalLevel=\"mpc\">\n")
            comm_data = ""
            if user[3] != "" or user[4] != "" or user[5] != "":
                comm_data = comm_data + '      <communicationData'
                if user[3] != "":
                    comm_data = comm_data + ' phoneNumber=\"' + user[3] + '\"'
                if user[4] != "":
                    comm_data = comm_data + ' mobile=\"' + user[4] + '\"'
                if user[5] != "":
                    comm_data = comm_data + ' email=\"' + user[5] + '\"'
                comm_data = comm_data + '/>\n'
                salesman_output.write(comm_data)
            salesman_output.write("    </parameter>\n")
            salesman_output.write("  </invocation>\n")
        salesman_output.write("</common:ServiceInvocationCollection>")
    return out_file_path


def crosscheck_users(found_users_file):
    # compare users in db and file and return users found in file but not in db
    global found_users_db
    new_users = []

    uid, first_name, name = zip(*found_users_file)
    found_users_db = [x.lower() for x in found_users_db]
    for u in uid:
        if u.lower() not in found_users_db:
            new_users.append(u)
    return new_users


def get_db_users(tenant, button):
    # pull users from db and update found_users_db
    global load_done
    global found_users_db

    button.config(state=DISABLED)

    salesman_stmt = "SELECT DEALERDIRECTORYUID from icon.PARTNER_PARTNER WHERE TENANTID=\'" + tenant + "\' AND PARTNERTYPE='salesman';"

    tmp_found_users_db = connection.execute_query(salesman_stmt)

    for entry in tmp_found_users_db:
        found_users_db.append(entry[0])

    load_done.set(1)
    return None


def read_csv_file():
    # read csv file and return list with users
    found_users = []
    global input_file_path
    with open(input_file_path, 'rb') as input_file:
        input_reader = csv.reader(input_file, delimiter=";")
        for line in input_reader:
            found_users.append([line[0], line[1], line[2]])
    return found_users[1:]


def select_file():
    # ask user for filename and update global file path
    global input_file_path
    global file_set
    input_file_path = askopenfilename(initialdir=os.getcwd(), title="Choose the csv file",
                                      filetypes=[("CSV", "*.csv")])
    file_set.set(1)
    return None


def read_config():
    # read configuration data from file
    global config_file_path
    global tenants
    try:
        config = yaml.safe_load(open(config_file_path))
        tenants = config["tenants"]
    except IOError:
        tkMessageBox.showerror("I/O Error", "Error reading configuration file from %s" % config_file_path)
    return None


def connect_to_db(connection_button):
    global connection_established
    global connection

    connection_button.config(state=DISABLED)
    connection.show_db_data_window()

    connection_established.set(1)


def exit_function():
    global root
    connection.close_connection()
    os._exit(-1)


def main():
    # display main window and options
    global root
    global file_set
    global load_done
    global tenants

    root.protocol("WM_DELETE_WINDOW", exit_function)
    root.title("Check Salesman")
    file_label = Label(root, text="1. Select csv file containing dduser list.")
    file_label.grid(row=0, column=0, columnspan=2, padx=10, pady=10)
    select_button = Button(root, text="select", command=lambda: select_file())
    select_button.grid(row=0, column=2, padx=10, pady=10)
    root.wait_variable(file_set)

    select_button.config(state=DISABLED)
    found_users_file = read_csv_file()

    connection_label = Label(root, text="2. Click the connect button.")
    connection_label.grid(row=1, column=0, padx=10, pady=10)
    connection_button = Button(root, text="connect", command=lambda: connect_to_db(connection_button))
    connection_button.grid(row=1, column=2, padx=10, pady=10)
    root.wait_variable(connection_established)

    option_label = Label(root, text="3. Select the tenant where you want to crosscheck the salesman.")
    option_label.grid(row=2, column=0, columnspan=2, padx=10, pady=10)
    tenant_label = Label(root, text="tenant")
    tenant_label.grid(row=3, column=0, padx=10, pady=10)

    tenant = StringVar()
    tenant_dropdown = OptionMenu(root, tenant, *sorted(tenants))
    tenant_dropdown.grid(row=4, column=0, padx=10, pady=10)

    exec_label = Label(root, text="3. Press the load button to get the users from the database.")
    exec_label.grid(row=5, column=0, columnspan=2, padx=10, pady=10)
    exec_button = Button(root, text="load", command=lambda: get_db_users(tenant.get(), exec_button))
    exec_button.grid(row=5, column=2, padx=10, pady=10)
    root.wait_variable(load_done)

    new_users = crosscheck_users(found_users_file)
    if len(new_users) == 0:
        if tkMessageBox.askyesno("Sucess",
                                 "All users from the file are present in the database on the server you are connected to. Do you want to quit?"):
            connection.close_connection()
            exit(0)
    else:
        if tkMessageBox.askyesno("New users found",
                                 "There were %d users in the file but not in the database. Do you want to create a XML-file with the users?" % len(
                                     new_users)):
            user_data = query_user_data(new_users)
            salesman_out_file = create_salesman_xml_files(user_data, tenant.get())
            dealer_out_file, user_errors, org_errors = create_dealer_xml_files(user_data, tenant.get())
            errors = write_error_files(user_errors, org_errors)
            if errors:
                if tkMessageBox.askyesno("Partial success",
                                         "The XML-Files are located at %s. The file containing error messages is located at %s. Please fill out the sheet and use it with the script diff_salesman.py" % (
                                                 os.path.dirname(salesman_out_file), errors)):
                    connection.close_connection()
                    exit(0)
            else:
                if tkMessageBox.askyesno("Success",
                                         "The XML-Files are located at %s Do you want to quit?" % os.path.dirname(
                                             salesman_out_file)):
                    connection.close_connection()
                    exit(0)
        else:
            connection.close_connection()
            exit(0)
    root.mainloop()
    connection.close_connection()
    exit(0)


if __name__ == '__main__':
    read_config()
    main()
