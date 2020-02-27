#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: Excel_To_UTF-8_CSV
# Author: Benjamin Schmelcher
# Date: 20181214
# Description: Reads an excel sheet and provides utf-8 csv
# Version: 1.0.0
import csv
import os
import tkMessageBox
from Tkinter import Tk, IntVar, DISABLED
from tkFileDialog import askopenfilename
from ttk import Label, Button

from openpyxl import load_workbook

root = Tk()
input_file_path = ""
file_chosen = IntVar()
excel_rows = []
file_created = IntVar()

def exit_function():
    os._exit(-1)


def select_input():
    global input_file_path
    input_file_path = askopenfilename(title="Select the excel file",
                                      filetypes=[("Excel file", "*.xlsx")])
    file_chosen.set(1)

def create_csv_file():
    book = load_workbook(input_file_path)

    sheet = book.get_sheet_by_name(book.get_sheet_names()[0])

    for row in sheet.iter_rows():
        tmp = []
        for cell in row:
            if type(cell.value) is unicode:
                tmp.append(cell.value.encode("utf-8"))
            else:
                tmp.append(cell.value)
        excel_rows.append(tmp)

    with open(os.path.dirname(input_file_path)+"/output.csv", "wb") as output:
        out_writer = csv.writer(output, delimiter=";", lineterminator="\n")
        for row in excel_rows:
            out_writer.writerow(row)







def main():
    root.protocol("WM_DELETE_WINDOW", exit_function)
    inputLabel = Label(root, text="Please select the input .xlsx file")
    inputLabel.grid(row=0, column=0, padx=10, pady=10)
    inputButton = Button(root, text="select", command=lambda: select_input())
    inputButton.grid(row=0, column=1, padx=10, pady=10)
    root.wait_variable(file_chosen)
    inputButton.config(state=DISABLED)

    createLabel = Label(root, text="Creating file. Please wait...")
    createLabel.grid(row=1, column=0, padx=10, pady=10)
    create_csv_file()
    #root.wait_variable(file_created)

    if tkMessageBox.askyesno("File created", "The file has been created in {}. Do you want to quit?".format(os.path.dirname(input_file_path))):
        exit(0)






    root.mainloop()

    exit(0)


if __name__ == '__main__':
    main()
