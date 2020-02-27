#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: gems_remove_tenants
# Author: Benjamin Schmelcher
# Date: 2019-01-11
# Description: creates protocol which tenants of gems upload files use icon or deletes foreign tenants based on protocol
# Version: 1.0
import csv
import os
import time
import tkMessageBox
from Tkconstants import DISABLED
from Tkinter import Tk, Toplevel, IntVar
from tkFileDialog import askopenfilenames, askopenfilename
from ttk import Label, Button

from openpyxl import Workbook, load_workbook

root = Tk()
out_path = ""
upload_paths = ""


def protocol(protocol_button):
    protocol_button.config(state=DISABLED)

    allowed_orgs = []
    gems_orgs = []

    orgs_selected = IntVar()
    upload_selected = IntVar()
    create_done = IntVar()

    def select_orgs():
        select_orgs_button.config(state=DISABLED)
        input_paths = askopenfilenames(parent=top, title="Select dd_orgs files", filetypes=[("csv file", "*.csv")])
        for path in input_paths:
            with open(path, 'rb') as input:
                reader = csv.reader(input, delimiter=";")
                for idx, row in enumerate(reader):
                    if idx > 0 and row[5].encode("utf-8").lower() == "dealer":
                        allowed_orgs.append(row[6].encode("utf-8"))
        orgs_selected.set(1)

    def select_upload_files():
        global out_path
        global upload_paths

        select_upload_files_button.config(state=DISABLED)
        upload_paths = askopenfilenames(parent=top, title="Select GEMS upload files", filetypes=[("csv file", "*.csv")])

        out_path = os.path.dirname(upload_paths[0]) + "/Missing_organisations_{}.xlsx".format(time.strftime("%Y_%m_%d"))

        for file in upload_paths:
            with open(file, 'rb') as input:
                reader = csv.reader(input, delimiter=";")
                for idx, row in enumerate(reader):
                    if idx > 0 and row[2].encode("utf-8") not in gems_orgs:
                        gems_orgs.append(row[2].encode("utf-8"))
        upload_selected.set(1)

    def create_protocol():
        create_button.config(state=DISABLED)
        book = Workbook()
        sheet = book.active
        sheet.title = "Missing organisations"
        sheet.append(["Organisation"])

        for gems_org in gems_orgs:
            if gems_org not in allowed_orgs:
                sheet.append([gems_org])

        try:
            book.save(out_path)
        except:
            tkMessageBox.showerror("Error", "Error writing to file.\nPlease make sure you have not opened the protocol.")
            exit(-1)
        create_done.set(1)

    top = Toplevel()
    top.title("Create Protocol")

    select_orgs_label = Label(top, text="1. Select dd_orgs files for tenants which shall be included")
    select_orgs_label.grid(row=0, column=0, padx=10, pady=10)
    select_orgs_button = Button(top, text="select", command=lambda: select_orgs())
    select_orgs_button.grid(row=0, column=1, padx=10, pady=10)
    top.wait_variable(orgs_selected)

    select_upload_files_label = Label(top, text="2. Select gems upload files")
    select_upload_files_label.grid(row=1, column=0, padx=10, pady=10)
    select_upload_files_button = Button(top, text="select", command=lambda: select_upload_files())
    select_upload_files_button.grid(row=1, column=1, padx=10, pady=10)
    top.wait_variable(upload_selected)

    create_label = Label(top, text="3. Press the create button")
    create_label.grid(row=2, column=0, padx=10, pady=10)
    create_button = Button(top, text="create", command=lambda: create_protocol())
    create_button.grid(row=2, column=1, padx=10, pady=10)
    top.wait_variable(create_done)
    tkMessageBox.showinfo("Success", "The protocol was created successfully under {}".format(out_path))

    exit(0)


def remove(remove_button):
    remove_button.config(state=DISABLED)

    protocol_selected = IntVar()
    upload_selected = IntVar()
    parse_done = IntVar()

    protocol_list = []
    gems_files = []

    def select_protocol():
        select_protocol_button.config(state=DISABLED)

        input_path = askopenfilename(parent=top, title="Select Missing_organisations protocol",
                                     filetypes=[("Excel", "*.xlsx")])
        book = load_workbook(input_path)
        sheet = book.get_sheet_by_name(book.get_sheet_names()[0])
        for idx, row in enumerate(sheet.iter_rows()):
            if idx > 0:
                protocol_list.append(row[0].value)
        protocol_selected.set(1)

    def select_upload_files():
        global out_path
        global upload_paths

        select_upload_files_button.config(state=DISABLED)

        upload_paths = askopenfilenames(parent=top, title="Select GEMS upload files", filetypes=[("csv file", "*.csv")])
        out_path = os.path.dirname(upload_paths[0]) + "/updated_gems_files"
        try:
            os.stat(out_path)
        except:
            os.mkdir(out_path)

        upload_paths = sorted(upload_paths)

        for file in upload_paths:
            with open(file, 'rb') as inputfile:
                reader = csv.reader(inputfile, delimiter=";")
                templist = []
                for row in reader:
                    encodedlist = []
                    for entry in row:
                        encodedlist.append(entry.encode("utf-8"))
                    templist.append(encodedlist)
                gems_files.append(templist)
        upload_selected.set(1)

    def parse_upload_files():
        parse_button.config(state=DISABLED)

        global out_path
        global upload_paths

        for idx, file_list in enumerate(gems_files):
            try:
                with open(out_path + "/" + os.path.basename(upload_paths[idx]), 'wb') as outputfile:
                    writer = csv.writer(outputfile, delimiter=";")
                    for row in file_list:
                        if row[2] not in protocol_list:
                            templist = []
                            for entry in row:
                                templist.append(entry.decode("utf-8"))
                            writer.writerow(templist)
            except:
                tkMessageBox.showerror("Error", "Error writing to file.\nPlease make sure you have not opened any of "
                                                "the output files.")
                exit(-1)
        parse_done.set(1)

    top = Toplevel()

    select_protocol_label = Label(top, text="1. Select the Missing_organisations protocol")
    select_protocol_label.grid(row=0, column=0, padx=10, pady=10)
    select_protocol_button = Button(top, text="select", command=lambda: select_protocol())
    select_protocol_button.grid(row=0, column=1, padx=10, pady=10)
    top.wait_variable(protocol_selected)

    select_upload_files_label = Label(top, text="2. Select gems upload files")
    select_upload_files_label.grid(row=1, column=0, padx=10, pady=10)
    select_upload_files_button = Button(top, text="select", command=lambda: select_upload_files())
    select_upload_files_button.grid(row=1, column=1, padx=10, pady=10)
    top.wait_variable(upload_selected)

    parse_label = Label(top, text="3. Press the parse button")
    parse_label.grid(row=2, column=0, padx=10, pady=10)
    parse_button = Button(top, text="parse", command=lambda: parse_upload_files())
    parse_button.grid(row=2, column=1, padx=10, pady=10)
    top.wait_variable(parse_done)

    tkMessageBox.showinfo("Success", "The outputfile(s) were created successfully under {}".format(out_path))

    exit(0)


def main():
    root.title("GEMS remove tenants")

    selection_label = Label(root,
                            text="Create protocol for organisations not found\nor remove wrong organisations from gems files?")
    selection_label.grid(row=0, column=0, padx=10, pady=10, columnspan=2)

    protocol_button = Button(root, text="protocol", command=lambda: protocol(protocol_button))
    protocol_button.grid(row=1, column=0, padx=10, pady=10)

    remove_button = Button(root, text="remove", command=lambda: remove(remove_button))
    remove_button.grid(row=1, column=1, padx=10, pady=10)

    root.mainloop()

    exit(0)


if __name__ == '__main__':
    main()
