#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: gems_user_to_role
# Author: Leon Kiefer
# Date: 2019-08-01
# Description: converts GEMS user rights to excel sheet
# Version: v1.0


import csv
import tkMessageBox
from Tkinter import *
from tkFileDialog import askopenfilenames
from ttk import *
import time
import codecs
from openpyxl import *
from openpyxl.styles import *


root = Tk()
files_done = IntVar()
read_done = IntVar()
roles = []
user_role_outlet_tuples = []
user_tenant_tuples = []

file_path = ""


tenant_map = {"D-Outlet.1452757": "MBBEL","D-Outlet.1452799": "MBF", "D-Outlet.1452850": "MBP","D-Outlet.1453005": "MBBR","D-Outlet.XY03674007": "BMBS", "D-Outlet.7101": "MBBEL","D-Outlet.XY03665324": "MBBEL", "D-Outlet.1452741": "MBD", "D-Outlet.10334": "MBBR", "D-Outlet.1453037": "BMBS",
              "D - Outlet.XY03653402":"MBBR", "D-Outlet.7088":"MBBEL", "D-Outlet.1452815": "MBUK", "D-Outlet.PT01000019": "MBP", "D-Outlet.5310004855": "MBF", "D-Outlet.XY03655166":"BMBS","TEAM30246970":"BMBS", "TEAM66192876": "BMBS","D-Country.CN": "BMBS","E-Country.CN": "BMBS", "TEAM08897546": "BMBS",
              "TEAM70051931": "BMBS", "TEAM08385479": "BMBS", "D-Outlet.XY03642853": "MBBEL", "D-Outlet.XY03662802": "BMBS", "D-Outlet.XY03654046": "BMBS", "D-Outlet.XY03666093": "BMBS", "D-Outlet.XY03679884":"MBP", "D-Outlet.1452800": "MBF", "D-Outlet.XY03654805":"BMBS", "D-Outlet.XY03664187":"BMBS",
              "E-1854":"BMBS", "D-Outlet.XY03664183": "BMBS", "TEAM58170224":"BMBS", "D-Outlet.XY03647465": "BMBS", "D-Outlet.7071": "MBBEL", "E-1211": "MBF", "D-Outlet.XY03691415": "MBBR", "D-Outlet.5310001960": "MBF","D-Outlet.XY03673399":"MBF", "D-Outlet.7037":"MBBEL", "D-Outlet.PT01000045":"MBP",
              "D-Outlet.PT01000027":"MBP", "D-Outlet.7020":"MBBEL", "D-Outlet.DE00000153":"MBBEL", "D-Outlet.XY03662930":"MBBEL", "D-Outlet.XY03676687": "BMBS", "D-Outlet.XY03660166": "BMBS", "D-Outlet.XY03665620": "BMBS", "D-Outlet.XY03666912":"BMBS", "D-Outlet.XY03663262":"BMBS", "D-Outlet.XY03666937":"BMBS",
              "D-Outlet.XY03674552": "BMBS", "D-Outlet.XY03660245": "BMBS", "D-Outlet.XY03645928": "MBBEL", "D-Outlet.7091":"MBBEL", "D-Outlet.7066":"MBBEL", "D-Outlet.XY03643771": "MBF", "D-Outlet.5310011199": "MBF", "D-Outlet.5310011168":"MBF", "D-Outlet.XY03642825": "MBBEL", "D-Outlet.DE03641752": "MBBEL",
              "D-Outlet.7086":"MBBEL", "D-Outlet.7096": "MBBEL", "D-Outlet.7112": "MBBEL", "D-Outlet.XY03665688": "MBBEL", "D-Outlet.5310004461": "MBF", "D-Outlet.1452868": "MBE", "D-Outlet.XY03667288": "MBBR", "D-Outlet.5310008745": "MBF", "D-Outlet.1452763": "MBDan", "D-Outlet.1452811": "MBGR",
              "D-Outlet.1452825":"MBI", "D-Outlet.1452835": "MBNL", "D-Outlet.1452836":"MBNL", "D-Outlet.1452841":"MBOE", "D-Outlet.1452848":"MBPL", "D-Outlet.1452854": "MBSve", "D-Outlet.1452864": "MBCZ", "D-Outlet.CH00000209":"MBCH", "D-Outlet.5310008767":"MBF",
              "D-Outlet.XY03653414":"MBF", "D-Outlet.5310008772":"MBF", "D-Outlet.10519": "MBP", "D-Outlet.10526":"MBBR", "D-Outlet.10529":"MBBR", "D-Outlet.10538": "MBBR", "D-Outlet.10539": "MBBR", "D-Outlet.10556": "MBBR", "D-Outlet.XY03648445": "MBBR", "D-Outlet.XY03648452": "MBBR",
              "D-Outlet.XY03648454": "MBBR", "D-Outlet.XY03655147":"MBBR", "D-Outlet.XY03660522":"MBBR", "D-Outlet.XY03660531":"MBBR", "D-Outlet.XY03660532": "MBBR", "D-Outlet.XY03660533":"MBBR", "D-Outlet.XY03660534": "MBBR", "D-Outlet.XY03660535": "MBBR", "D-Outlet.XY03660537": "MBBR", "D-Outlet.XY03660538": "MBBR",
              "D-Outlet.XY03660539": "MBBR", "D-Outlet.XY03661361":"MBBR", "D-Outlet.XY03661401":"MBBR", "D-Outlet.XY03662279":"MBBR", "D-Outlet.XY03662894":"MBBR", "D-Outlet.XY03662897":"MBBR", "D-Outlet.XY03663375":"MBBR", "D-Outlet.XY03664045":"MBBR", "D-Outlet.XY03665260":"MBBR", "D-Outlet.XY03665769":"MBBR",
              "D-Outlet.XY03666130":"MBBR", "D-Outlet.XY03666647":"MBBR", "D-Outlet.XY03666833":"MBBR", "D-Outlet.XY03667102":"MBBR", "D-Outlet.XY03667378":"MBBR", "D-Outlet.XY03668368":"MBBR", "D-Outlet.XY03668460":"MBBR", "D-Outlet.XY03668795":"MBBR", "D-Outlet.XY03668942": "MBBR", "D-Outlet.XY03669623": "MBBR",
              "D-Outlet.XY03669869":"MBBR", "D-Outlet.XY03669870":"MBBR", "D-Outlet.XY03669944": "MBBR", "D-Outlet.XY03670722":"MBBR", "D-Outlet.XY03671353": "MBBR","D-Outlet.XY03671365":"MBBR", "D-Outlet.XY03671911":"MBBR", "D-Outlet.XY03672014":"MBBR", "D-Outlet.XY03672015":"MBBR", "D-Outlet.XY03672049":"MBBR",
              "D-Outlet.XY03673309":"MBBR", "D-Outlet.XY03673701":"MBBR", "D-Outlet.XY03673884":"MBBR", "D-Outlet.XY03691826":"MBBR", "D-Outlet.5310008725":"MBF","D-Outlet.5310005231":"MBF", "D-Outlet.XY03647566":"MBF", "D-Outlet.XY03665848":"MBF", "D-Outlet.5310008836":"MBF", "D-Outlet.5310000491":"MBF",
              "D-Outlet.Z531910024":"MBF", "D-Outlet.XY03646726":"MBF", "D-Outlet.1452750":"MBD","D-Outlet.7028": "MBBEL", "D-Outlet.7122":"MBBEL", "D-Outlet.7023":"MBBEL", "D-Outlet.1452942": "MBSA", "D-Outlet.XY03665391":"MBD", "D-Outlet.5310001811":"MBF", "D-Outlet.5310000690": "MBF",
              "D-Outlet.5310005080": "MBF", "D-Outlet.5310011062": "MBF", "D-Outlet.XY03642192":"Monaco", "D-Outlet.XY03662869":"MBF", "D-Outlet.XY03673401":"MBF", "D-Outlet.XY03655393":"MBF", "D-Outlet.XY03677235":"MBBEL", "D-Outlet.10351":"MBBR", "D-Outlet.XY03664791":"MBBR", "D-Outlet.7106":"MBBEL",
              "D-Outlet.7108":"MBBEL", "D-Outlet.7110": "MBBEL", "D-Outlet.XY03643816":"MBBEL", "D-Outlet.1452809":"MBF", "D-Outlet.XY03654040":"MBF", "D-Outlet.5310003467": "MBF", "D-Outlet.XY03651025":"MBBEL", "D-Outlet.PT01000028": "MBP", "C-Outlet.XY03668447": "MBF", "D-Outlet.5310003600":"MBF",
              "D-Outlet.XY03665181":"MBF", "D-Outlet.10711":"BMBS", "D-Outlet.XY03675225":"MBF", "D-Outlet.10305":"MBBR", "D-Outlet.XY03674918": "MBF", "D-Outlet.5310000614": "MBF", "D-Outlet.XY03658768":"MBF", "D-Outlet.5310003881": "MBF", "D-Outlet.XY03651562": "MBP", "C-Outlet.XY03681338":"MBNL",
              "D-Outlet.8441":"MBNL", "D-Outlet.XY03661420":"MBNL", "D-Outlet.XY03661476":"MBNL", "D-Outlet.XY03661475": "MBNL", "D-Outlet.5310004256":"MBF", "D-Outlet.XY03653402": "MBBR"}



'''
rauch_test_file = "Rauch_Test.csv"

with open(rauch_test_file,'rb') as input_file:
    input_reader = csv.reader(input_file,delimiter =";")
    for i, line in enumerate(input_reader):
        try:
            #print(tenant_map[line[0]])
            result = tenant_map[line[0]]
            print(result)
        except:
            print("fail", line)'''



def sort_first(val):
    return val[0]

def select_files():
    # select one or multiple gems files
    global gems_input_path
    global files_done
    input_path = askopenfilenames(title="select one or more gems files", filetypes=[("csv file", "*.csv")])
    gems_input_path = list(input_path)
    files_done.set(1)
    return None




def read_csv():

    for file in gems_input_path:
        with open(file, 'rb') as input_file:
            input_reader = csv.reader(input_file, delimiter=";")
            for i, line in enumerate(input_reader):
                # [user_id, user_surname, user_name, user_mail, user_country, user_orgs, user_roles]
                temp_user = line[0]
                temp_user_surname = codecs.decode(line[7], 'cp1252')
                temp_user_first_name = codecs.decode(line[8], 'cp1252')
                temp_user_mail = codecs.decode(line[9], 'cp1252')
                temp_user_country = codecs.decode(line[26], 'cp1252')
                temp_dealer_name = codecs.decode(line[13], 'ISO-8859-1')
                temp_tech_org_id = line[2]
                temp_role = re.sub("ICON_[0-9]_","",line[6])

                try:
                    temp_tenant = tenant_map[temp_tech_org_id]
                except:
                    temp_tenant = "fail"

                temp_user_tenant_tuple = (line[0], temp_tenant)
                temp_user_role_outlet_tuple = (temp_user, temp_user_surname,  temp_user_first_name, temp_user_mail, temp_user_country, temp_dealer_name, temp_tenant, [temp_role])

                if temp_user_tenant_tuple not in user_tenant_tuples and temp_user != "User-Id":
                    user_role_outlet_tuples.append(temp_user_role_outlet_tuple)
                    user_tenant_tuples.append(temp_user_tenant_tuple)

                elif temp_user_tenant_tuple in user_tenant_tuples:
                    [item[7].append(temp_role) for item in user_role_outlet_tuples if item[0] == temp_user and item[6] == temp_tenant and temp_role not in item[7]]

                #if temp_user not in user:
                #    user.append(temp_user)

                if temp_role not in roles and temp_role != "Name":
                    roles.append(temp_role)

    roles.sort()
    user_role_outlet_tuples.sort(key = sort_first)
    read_done.set(1)
   # print([item for item in user_role_outlet_tuples if item[0] == "MATHST"])

'''
for user in user_role_tuples[0]:
    if temp_user !=
if temp_user not in user_role_tuples



if [item for item in user_role_tuples if item[0] == temp_user]:
    temp_tuple = [item for item in user_role_tuples if item[0] == temp_user]
    if temp_role not in temp_tuple[1]:
        temp_tuple[1].append(temp_role)
if temp_user in temp_user_role_tuple[0] and temp_role not in temp_user_role_tuple[1]:
    temp_user_role_tuple[1].append(temp_role)
user_role_tuples.append(temp_user_role_tuple)
#temp_role = line[1][7:]

)'''
   # print(user)
    #print(roles)
    #print(user_role_tuples)



def write_excel_file():
    start = time.time()
    file_path = os.path.dirname(gems_input_path[0]) + "/gems_user_rights_" + time.strftime("%Y_%m_%d") + ".xlsx"
    global found_users
    book = Workbook()
    sheet = book.active
    header = ["UserID", "Surname", "First Name", "E-Mail", "Country", "Dealer Name", "Tenant"]
    for item in roles:
        header.append(item)
    sheet.append(header)
    # [user_id, user_surname, user_name, user_mail, user_country, user_orgs, user_roles]
    for idxTup, tuple in enumerate(user_role_outlet_tuples,1):
        sheet.cell(row=(idxTup +1),column = 1).value = tuple[0]
        sheet.cell(row = idxTup +1, column = 2).value = tuple[1]
        sheet.cell(row=idxTup + 1, column=3).value = tuple[2]
        sheet.cell(row=idxTup + 1, column=4).value = tuple[3]
        sheet.cell(row=idxTup + 1, column=5).value = tuple[4]
        sheet.cell(row=idxTup + 1, column=6).value = tuple[5]
        sheet.cell(row=idxTup + 1, column=7).value = tuple[6]
        idxCol = 1

        while idxCol <= sheet.max_column:
            temp_cell = sheet.cell(row=1, column = idxCol)
            if temp_cell.value in tuple[7]:
                sheet.cell(row=(idxTup + 1), column=idxCol).value = "x"
            idxCol += 1


    book.save(file_path)
    end = time.time()
    print(end - start)
    start1 = time.time()
    style_excel_file(book, file_path)
    end1 = time.time()
    print(end1- start1)




def style_excel_file(book, filepath):
    global tenant_list
    global right_list
    try:
        blue = "004d94ff"
        orange = "00ffb366"
        light_blue = "0066c2ff"
        light_green = "0078e797"
        light_yellow = "00ffbf00"
        #book = load_workbook(filepath)
        sheet = book.active
        # first column
        sheet.cell(row=1, column=1).fill = PatternFill(patternType='solid', fgColor=blue, )
        for j in range(2, sheet.max_row + 1):
            sheet.cell(row=j, column=1).font = Font(bold=True)
            sheet.cell(row=j, column=1).fill = PatternFill(patternType='solid', fgColor=blue)
        # user data
        for i in range(2, 5):
            for j in range(1, sheet.max_row + 1):
                sheet.cell(row=j, column=i).fill = PatternFill(patternType='solid', fgColor=orange)
        # language
        for i in range(5, 6):
            for j in range(1, sheet.max_row + 1):
                sheet.cell(row=j, column=i).fill = PatternFill(patternType='solid', fgColor=light_blue)

        # Dealer Name
        for j in range(1, sheet.max_row + 1):
            sheet.cell(row=j, column=6).fill = PatternFill(patternType='solid', fgColor=light_green)

        # tenants
        tenant_end = 7 # second number is number of different tenants+1
        #for i in range(7, tenant_end):
        for j in range(1, sheet.max_row + 1):
            sheet.cell(row=j, column=7).fill = PatternFill(patternType='solid', fgColor=light_green)
        # entitlements
        for i in range(tenant_end +1, sheet.max_column + 1):
            for j in range(1, sheet.max_row + 1):
                sheet.cell(row=j, column=i).fill = PatternFill(patternType='solid', fgColor=light_yellow)
        # borders
        for i in range(1, sheet.max_column + 1):
            for j in range(1, sheet.max_row + 1):
                sheet.cell(row=j, column=i).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                            top=Side(style='thin'), bottom=Side(style='thin'))
                sheet.cell(row=j, column=i).alignment = Alignment(horizontal="center", vertical="center")
        # header
        for i in range(1, sheet.max_column + 1):
            tmp = sheet.cell(row=1, column=i)
            tmp.font = Font(bold=True)
            if i > 7:
                tmp.alignment = Alignment(textRotation=90, horizontal="center", vertical="center", wrap_text=True)
        # size columns
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 21
        sheet.column_dimensions['C'].width = 18
        sheet.column_dimensions['D'].width = 40
        sheet.column_dimensions['F'].width = 65
        sheet.column_dimensions['G'].width = 20
        book.save(filepath)
    except IOError:
        tkMessageBox.showerror("IO/Error", "Outputfile %s not found for styling!" % filepath)
    return None




def main():
    global root
    global files_done
    global read_done
    global error_log
    global connection_set
    global environment
    root.title("GEMS user rights")



    Label(root, text="1. Select the gems file").grid(row=4, column=0, padx=10, pady=10)
    b1 = Button(root, text="select", command=lambda: select_files())
    b1.grid(row=4, column=2, padx=10, pady=10)
    root.wait_variable(files_done)
    b1.config(state=DISABLED)
    Label(root, text="2. Press the convert button").grid(row=5, column=0, padx=10, pady=10)
    b2 = Button(root, text="convert", command=lambda: read_csv())
    b2.grid(row=5, column=2, padx=10, pady=10)
    root.wait_variable(read_done)
    b2.config(state=DISABLED)
    Label(root, text="Writing Excel Output. This may take a bit").grid(row=6, column=0, padx=10, pady=10)
    root.update()
    write_excel_file()
    if tkMessageBox.askokcancel("Success",
                                "The Excel file has been successfully created in. Do you want to quit?"):

        exit(0)


if __name__ == '__main__':
    main()

