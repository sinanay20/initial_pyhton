#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Name: Dealers xlsx to xml
# Author: Benjamin Schmelcher
# Date: 10.09.2018
# Description: Provides a dealers.xml file based on a error file (format see documentation)
# Version: 0.0.1

import csv
import os
import tkMessageBox
from Tkinter import Tk, IntVar
from tkFileDialog import askopenfilename
from ttk import Label, Button

from openpyxl import load_workbook

#output_file_name = "./output/Dealers.xml"
#xlsx_file_name = "./Copy of errors_2018_09_06 (002).xlsx"
#error_file = "./output/DealerErrors.xml"
salesmen = []
dealers = []

root = Tk()
file_selected = IntVar()
xlsx_file_name = ""
output_file_name = ""
error_file = ""


def create_xml_file():
    global output_file_name
    global error_file

    try:
        output_writer = open(output_file_name, 'w')
    except:
        tkMessageBox.showerror("I/O Error",
                               "Can not open file" + os.path.dirname(output_file_name) + "/" + output_file_name + "for writing")

        exit(1)


    #create Header in file
    output_writer.write(
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n"
        "<common:ServiceInvocationCollection xmlns:common=\"http://common.icon.daimler.com/il\" "
        "xmlns:partner_pl=\"http://partner.icon.daimler.com/pl\" "
        "xmlns:mdsd_sl=\"http://system.mdsd.ibm.com/sl\" "
        "xmlns:xsi=\"http://www.w3.org/2001/XMLSchema-instance\" "
        "xmlns:xsd=\"http://www.w3.org/2001/XMLSchema\">\n"
        
        "<!--Related to CIM: 20140811_CIM_EDF_OrganisationalPerson(dealer)_Mig_BEL_WavePreInt4_iter1_v1.0.xlsx-->\n"
        "<!--Related to Masterdata:5.11-->\n"
        "<!--Source-database: db3.S415VM779.tst-->\n"
        
        "  <executionSettings xsi:type=\"mdsd_sl:ExecutionSettingsType\" "
        "dateTime=\"2018-09-06T00:00:00\" userId=\"ICON_tec_EDF01\" "
        "tenantId=\"81930CN\" causation=\"migration\" "
        "additionalInformation1=\"1\" "
        "issueThreshold=\"error\"/>\n"
        )

        #über dealer iterieren
    for dealer in dealers:
        contains_salesman = None
        for salesman in salesmen:
            if(salesman[6]==dealer[28]):
                contains_salesman = True

        if(contains_salesman):
            output_writer.write(
            "  <invocation operation=\"updateOrganisationalPerson\">\n"
            "    <parameter xsi:type=\"partner_pl:OrganisationalPersonType\" "
            "externalId=\"""\" "
            "sourceSystem=\"migration\" "
            "masterDataReleaseVersion=\"9\" "
            "partnerType=\"dealer\" "
            "state=\""+dealer[10]+"\" "
            "migrationDate=\"2018-09-06T00:00:00\" "
            
            "companyName=\""+dealer[21]+"\" "
            )

            if(dealer[27] != "<null>"):
                output_writer.write(
                "companyName2=\""+dealer[27]+"\" "
                )
            else:
                output_writer.write(
                "companyName2=\"\" "
                )

            output_writer.write(
            "companyInternal=\"no\" "
            "vatId=\"\" "
            "gssnOutletCompanyId=\"\" "
            "gssnOutletOutletId=\""+dealer[28]+"\">\n"
    
            "      <communicationData phoneNumber=\"\" faxNumber=\"\"/>\n"
    #TODO: city
            "      <legalAddress city=\"\" "
            "country=\"CN\" "
            "street=\"\" "
            "zipCode=\"\"/>\n"
            )



            #über personen iterieren
            for index, salesman in enumerate(salesmen):

                salesmen[index].append(False)
                if(salesman[6]==dealer[28]):
                    salesmen[index][7] = True
                    output_writer.write(
                    "        <contactPartnerAssignment internal=\"false\" "
                    "salesman=\"true\">\n"
                    "          <contactPerson xsi:type=\"partner_pl:PhysicalPersonType\" "
                    "externalId=\""+salesman[0]+"\" "
                    "sourceSystem=\"migration\"/>\n"
                    "        </contactPartnerAssignment>\n"
                    )
            #end
            output_writer.write(
            "    </parameter>\n"
            )
    output_writer.write(
    "</common:ServiceInvocationCollection >"
    )
    output_writer.close()

    try:
        error_output_writer = open(error_file, 'w')
        error_output_writer.write(
            "Salesmen without Dealer:\n\n")
        for index, element in enumerate(salesmen):
            if (element[7] != True):
                error_output_writer.write(
                    "Index: " + str(index) + "       UID: " + element[0] + "\n"
                )

    except:
        tkMessageBox.showerror("I/O Error", "Can not open file" + os.path.dirname(error_file) + "/" + error_file + "for writing")


def read_xlsx_file():
    global xlsx_file_name
    global salesmen
    book = load_workbook(xlsx_file_name)
    sheet = book.get_sheet_by_name("user data")

    for index, row in enumerate(sheet.iter_rows()):
        tempsalesman = []
        if index>0:
            for cell in row:
                tempsalesman.append(cell.value)
            salesmen.append(tempsalesman)

    sheet = book.get_sheet_by_name("DB Dealer")

    for index, row in enumerate(sheet.iter_rows()):
        tempdealer = []
        if index>0:
            for cell in row:
                tempdealer.append(cell.value)
            dealers.append(tempdealer)


def select_error_file():
    global root
    global xlsx_file_name
    global output_file_name
    global error_file
    xlsx_file_name = askopenfilename(parent=root, title="Select error file", filetypes=[("Excel", "*.xlsx")])
    output_file_name = os.path.dirname(xlsx_file_name)+"/Dealers.xml"
    error_file = os.path.dirname(xlsx_file_name)+"/Dealers_errors.xml"

    file_selected.set(1)



def main():
    global root
    selectLabel = Label(root, text="1. Select the errorfile containing the information to create the Dealers xml file.")
    selectLabel.grid(row=0, column=0, padx=10, pady=10)
    selectButton = Button(root, text="select", command=lambda: select_error_file())
    selectButton.grid(row=0, column=1, padx=10, pady=10)

    root.wait_variable(file_selected)


    read_xlsx_file()
    create_xml_file()

    tkMessageBox.showinfo(title="Success",
                          message="Succes. The outputfile has been created in {}.".format(output_file_name))

    exit(0)





if __name__ == '__main__':
    main()